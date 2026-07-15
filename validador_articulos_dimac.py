#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VALIDADOR COMPLETO DE ACCESORIOS - Thermia / Dimac  (v final)
=============================================================
Cruza las 5 tablas de Geinfor + el Excel de seguimiento de José María y valida
los accesorios FW creados/modificados según todas las reglas acordadas.

ARCHIVOS DE ENTRADA:
  1) REPASO.xlsx  con hojas:
       - MAESTRO DE ARTICULOS   (llave: CODIGO_ARTICULO)
       - ARTICULOS TARIFABLES   (llave: ARTICULO)      -> PVP (tarifa 40)
       - ARTICULOS PROVEEDOR    (llave: ARTICULO)      -> coste/plazo/lote (PRINCIPAL=1)
       - ALMACEN ARTICULOS      (llave: ARTICULO)      -> stock mínimo
       - TABLA CARACTERISTICAS  (llave: ARTICULO)      -> imagen (caract. 0)
  2) JOSEMARIA.xlsx  hoja 'RELLENAR JOSE MARIA' (llave: ARTÍCULO) -> divisor, mínimo

USO:
    python validador_final.py REPASO.xlsx JOSEMARIA.xlsx [USUARIO] [FECHA_DESDE]
Sin usuario/fecha -> valida todos los FW del maestro.
"""
import sys
import re
import html as _html
import unicodedata
import pandas as pd

# Nombres legibles de cada campo para el informe visual
NOMBRES_CAMPO = {
    "FAMILIA": "Familia de producto",
    "CLASE": "Clase / serie",
    "DENOMINACION": "Denominación",
    "DENOMINACION_2": "Denominación 2 (acabado)",
    "MEDIDA": "Medida",
    "COEFICIENTE": "Coeficiente",
    "UNIDADES_ARTICULO": "Unidades del artículo",
    "GRUPO": "Grupo (acabado)",
    "CARACTERISTICA_0": "Imagen",
    "PROVEEDOR": "Proveedor principal",
    "PLAZO_ENTREGA": "Plazo de entrega",
    "LOTE_HABITUAL": "Lote habitual",
    "MARGEN": "Margen / PVP",
    "STOCK_MINIMO": "Stock mínimo",
    "ESTADO": "Estado",
    "TIPO": "Tipo",
    "UNIDAD_PROVEEDOR": "Unidad del proveedor",
    "UNIDAD_VENTA": "Unidad de venta",
    "UNIDAD/TIPO": "Unidad / Tipo",
    "EXCEL_JM": "Excel José María",
    "DIVISOR": "Divisor",
    "PVP": "PVP",
    "IVA_TARIFA": "IVA en tarifa 40",
    "IVA_PROVEEDOR": "IVA en artículo proveedor",
    "FORMULA_PROVEEDOR": "Fórmula en artículo proveedor",
    "TIPOARTMODULAR": "Tipo artículo modular",
    "PARTIDAARANCELARIA": "Partida arancelaria",
    "CONVERSION_M_KG": "Coeficiente m/kg",
    "GESTIONARVARIABLES": "Gestionar variables",
    "ESSUBCONTRATACIO": "Es subcontratación",
    "ESKIT_CHECK": "Es un artículo kit",
    "NOGESTIONASTOCK": "No gestiona inventario",
    "CALCULOCANTFORMULA": "Cálculo cantidad por fórmula",
    "ARTICULOBASE": "Artículo base",
}

# Reglas de campo fijo del maestro. COEFICIENTE y UNIDADES_ARTICULO se validan aparte
# (dependen del caso de unidades: barras, metro lineal, unitario).
REGLAS_MAESTRO = {}
GRUPO_POR_SUFIJO = {"04": 10, "03": 11, "01": 12, "06": 14, "05": 16}
# Texto que debe llevar DENOMINACION_2 según GRUPO del maestro
DENOM2_POR_GRUPO = {10: "NEGRO", 11: "BLANCO", 12: "PLATA", 14: "RAL STD", 16: "PVD"}
# Inverso: para artículos con código no estándar, el grupo se deriva de la DENOMINACION_2
_GRUPO_POR_DEN2 = {v: k for k, v in DENOM2_POR_GRUPO.items()}
_GRUPO_POR_DEN2["MARRÓN"] = 10   # alias de NEGRO (grupo 10)
_GRUPO_POR_DEN2["MARRON"] = 10  # alias sin tilde
# Artículos con grupo asignado manualmente (correcto, no se valida)
_GRUPO_WHITELIST = {"TAPON04", "MANUAL0004"}

# Prefijos de artículos con código de 11 dígitos (base 8 + letra dirección D/I + 2 sufijo acabado).
# Para estos, si el sufijo completo no se encuentra, se usan los 2 últimos caracteres.
_PREFIJOS_SUFIJO_ULTIMOS2 = ("PM4028",)

# Artículos A0063*: sufijos AM/BM/IM/NM = acabados PVD especiales → GRUPO 16.
# DEN2 contiene el nombre específico del acabado (ANTRACITA MATE, BRONCE MATE, etc.),
# no el valor estándar "PVD", por lo que la validación de DEN2 se omite para estos.
_PREFIJOS_SUFIJO_PVD = ("A0063",)
_SUFIJO_GRUPO_PVD = {"AM": 16, "BM": 16, "IM": 16, "NM": 16}
TARIFA_PVP = 40
# Tolerancia de margen híbrida: solo se marca si la diferencia supera AMBOS umbrales.
# Absorbe el band sistemático (~3-4%, donde el PVP real queda algo por encima del
# calculado = más margen, validado con José María en el artículo 072040) y el redondeo
# de céntimos, pero sigue detectando los outliers reales (>10-70%).
TOL_MARGEN_EUR = 0.05  # diferencia mínima en € para considerar incidencia
TOL_MARGEN_PCT = 5.0   # diferencia mínima en % para considerar incidencia

# Artículos que pueden (y deben) tener NOGESTIONASTOCK=1.
# Si un artículo tiene el flag y NO está aquí → error.
# Si está aquí y NO tiene el flag → error.
# Lista validada manualmente por Adrián (jun 2026).
_NOSTOCK_WHITELIST = {
    "A0002032","A0002033",
    "A0009101","A0009102","A0009103","A0009104","A0009105","A0009106",
    "A0009107","A0009108","A0009109","A0009110","A0009111","A0009152","A0009190",
    "A0409140",
    "A0ME4015","A0ME4017","A0ME4025",
    "ABONO","ACCSVARIOS","ACUENTA","ALQUILER",
    "C062ME01","C062ME02","C062ME06",
    "CANON RIBATECH","CANON TECNICAL2",
    "CHATARRA","COMISIONESVENTA",
    "COMPANEL","COMPATINENCA",
    "CPTCORTE","CPTGRUESO1","CPTGRUESO2","CPTGRUESO3","CPTGRUESO4",
    "CPTPA","CPTRAL","CPTTAM","CPTTAM2","CPTTAM3",
    "CPTVIDRIO1","CPTVIDRIO2","CPTVIDRIO3",
    "CURVA","EMBALAJE","ENSAMBLADO","ENVIO",
    "ER72ME01","ER72ME02","ER72ME10","ER72ME30",
    "FLETE","FLETE ARGELIA","FLETE CHILE","FLETE COLOMBIA","FLETE COSTA RICA",
    "FLETE CURAZAO","FLETE ECUADOR","FLETE EL SALVADOR","FLETE ESTADOS UNIDOS",
    "FLETE MEXICO","FLETE PANAMA","FLETE PERU","FLETE PUERTO RICO",
    "FLETE REPUBLICADOMINICANA","FLETE URUGUAY","FLETE VENEZUELA",
    "GASTOS DEV.","GR50ME10",
    "LACADO","LACADO EXPRESS","LOCAL",
    "MANIPULACION","MANIPULACION 2","MANOOBRA",
    "METAR78CE","METCF3100","METCR4000","METER5200",
    "MINIMO","MINIMOACS","MINIMOACS2","MINIMOACS3",
    "MKT00004","MLAP7816",
    "OTROS SERVICIOS PERU","PANEL ATINENCA","PORTES",
    "RAPPELCONSUMO",
    "SEGURO","SEGURO CHILE","SEGURO COLOMBIA","SEGURO COSTA RICA","SEGURO CURAZAO",
    "SEGURO ECUADOR","SEGURO EL SALVADOR","SEGURO ESTADOS UNIDOS","SEGURO MEXICO",
    "SEGURO PANAMA","SEGURO PERU","SEGURO PUERTO RICO","SEGURO REP.DOMINICANA",
    "SEGURO URUGUAY",
    "SUPLEMENTO 1","SUPLEMENTO 2","SUPLEMENTO 3","SUPLEMENTO 4",
    "VARIOEX","VARIOS","VENTANA",
}

# Artículos que no llevan imagen (característica 0) por naturaleza.
_SIN_IMAGEN_WHITELIST = {
    "SUPLEMENTO 1","SUPLEMENTO 2","SUPLEMENTO 3","SUPLEMENTO 4",
}

# Familia 105 (chapas y paneles): clases válidas y DENOMINACION_2 esperada por grupo (tipo 92)
_F105_CLASES_VALIDAS = {31, 54, 56}

# Artículos de familia 105 que SÍ pueden tener imagen (excepciones justificadas)
# XTA* = chapas termoacústicas especiales, XDECOR* = chapas decorativas
_F105_CON_IMAGEN = {
    "XTA3125103",
}
_F105_CON_IMAGEN_PREFIJOS = ("XTA", "XDECOR")

# Artículos de familia 105 tipo 1 que ya tienen acabado (grupo != 0 permitido)
_F105_TIPO1_CON_ACABADO = {
    "XTA3125103",
}

# Artículos de familia 105 clase 0 bajo presupuesto: sin precio fijo en proveedor ni tarifa
_F105_BAJO_PRESUPUESTO = {
    "PANEL",
}
_F105_DEN2_POR_GRUPO = {
    2: "BRONCE",
    5: "ANODIZADO ESPECIAL",
    6: "RAL ESTANDAR",
    7: "RAL ESPECIAL",
    8: "BICOLOR",
    9: "MADERA",
}

# Excepciones validadas por Adrián:
# Margen: artículos con margen acordado / pendiente de revisión — no generar error
_MARGEN_WHITELIST = {
    # Acuerdos comerciales específicos
    "A0009109", "A0009110",
    # Tornillos — margen bajo conocido y aceptado
    "A0003237", "A0003238", "A0003239",
    # Márgenes actuales mantenidos (revisión pendiente o decisión comercial)
    "A000372903", "A000372904",
    "A0403752", "AR902312",
    "CR402345", "CR402346",
    "ER722307", "ER722308", "ER722342", "ER72350", "ER722351",
    "CR623741",
    "ER720800", "ER720801", "ER720802", "ER720900", "ER723700",
    # GR552002/03/04 — pendiente de validación
    "GR552002", "GR552003", "GR552004",
}
# Margen: artículos FW clase 53 — rango válido 20-23%
_MARGEN_FW_CLASE53_MIN = 20.0
_MARGEN_FW_CLASE53_MAX = 23.0
# Margen: clases pendientes de ajuste de tarifa (no generar error mientras tanto)
_MARGEN_CLASES_PENDIENTES = {38, 39}
# Plazo entrega: reglas explícitas por proveedor (sobreescriben el habitual calculado)
# {proveedor: plazo_esperado} — para proveedores con un único plazo correcto
_PLAZO_FIJO_PROV = {
    407: 12,
}
# {proveedor: {ui: plazo}} — para proveedores con plazo según unidad interna
_PLAZO_PROV_POR_UI = {
    599: {1: 15, 4: 20},
}
# {proveedor: {clase: plazo}} — para proveedores con plazo según clase de artículo
_PLAZO_PROV_POR_CLASE = {
    783: {47: 30},
}
# Clases sin validación de plazo (pendiente de ajuste)
_PLAZO_CLASES_PENDIENTES = {38}
# TIPO: artículos tipo 92 con CATALOGOINTERIORES pero sin artículo base en estructura (por diseño)
_TIPO92_SIN_BASE_PREFIJOS = ("RA470", "RA471", "RA461", "A007")
_TIPO92_SIN_BASE_CODIGOS = {
    "ER52090006", "ER520901",
    "A0003157", "A0003175",
    "A0003731", "A0003734S", "A0003735", "A0003738S",
}
# TIPO: artículos cuya base en la estructura es el código completo (no cod[:8])
_TIPO92_BASE_FULLCODE = {"A0003120D", "A0003120I"}
# CLASE: artículos con clase validada manualmente (no coincide con serie JM pero es correcto)
_CLASE_WHITELIST = {"A0002010", "A0002012"}
# LOTE_ALMACEN: artículos donde lote almacén ≠ lote proveedor es correcto por acuerdo
_LOTE_ALM_WHITELIST = set()
# UNIDAD_PROVEEDOR: artículos donde unidad externa maestro ≠ unidad proveedor es correcto
_UNIDAD_PROV_WHITELIST = set()
# TIPO: artículos sin estructura con tipo correcto fuera de la norma (1/2/5/60)
_TIPO_SIN_ESTRUCTURA_OK = {
    "CAT00011", "CAT00000", "A000317501",
    "VARIOS", "A000373306", "ACCSVARIOS",
    "MINIMOACS", "MINIMOACS2", "MINIMOACS3", "ACS2", "ACS3",
    "CR4450",
}
_TIPO_SIN_ESTRUCTURA_PREFIJOS = ("MQ",)
# Coeficiente proveedor ≠ maestro pero justificado
_COEF_PROV_WHITELIST = {"C0404202", "CF221500"}
# Artículos con coeficiente maestro ≠ 1 por conversión de unidades (correcto por diseño)
_COEF_MAESTRO_WHITELIST = {"CF221500"}
# Artículos base de lacado: no deben estar en estado 70 (no son vendibles directamente)
_ESTADO_WHITELIST = {"A0003252", "A0003261"}
# Artículos con UI ≠ UE que NO necesitan variable 10 (estructura especial)
_ARTVAR_WHITELIST = {"CF2241500", "CF221500", "CHATARRA", "ENSAMBLADO", "MAKILA"}
# Artículos con tipo, clase y unidad interna validados manualmente (configuración especial)
_TIPO_CLASE_UI_WHITELIST = {"ENSAMBLADO"}
# Tipo 92 grupo 14 que no llevan cargo por mínimo (van sin MINIMOACS).
# Se comprueba por prefijo para cubrir todas las variantes (GR55200206, RA47006, etc.)
_ARTASOC_SIN_MINIMO_PREFIJOS = ("GR55", "RA470", "RA471")
# Artículos con configuración de asociados validada manualmente (cantidad/integración no estándar)
# Incluye variantes (A000323706 = variante de A0003237, etc.)
_ARTASOC_LINEAS_OK = {
    "A0003157",
    "A000315706",
    "A0003237", "A000323706",
    "A0003238", "A000323806",
    "A0003239", "A000323906",
    "CR445006",
}


def num(v):
    try:
        f = float(v)
    except (ValueError, TypeError):
        return None
    return None if f != f else f  # NaN (f != f) -> None


def ni(v):  # entero o None
    f = num(v)
    return int(f) if f is not None else None


# Conexión directa a Geinfor (IBM DB2). SOLO LECTURA. La base de Thermia/Dimac es la DSN
# GeinprodINGENRED (conecta sin credenciales estando en la red de la oficina). Esquema DB2ADMIN.
DSN_DB2 = "GeinprodINGENRED"

# Carpeta de red donde José María mantiene su Excel de seguimiento. El validador coge solo
# el más reciente que cumpla el patrón, así no hay que copiarlo a mano a la carpeta local.
JM_CARPETA = r"\\dimac2023\DATOS\DATOS DIMAC\USERS\INTERCAMBIO\RAUL\xavi\PROCES DE COMPRES ADRIA\EXTRUSORES\MATRICES EXTRUSIÓN"
JM_PATRON = "*Matrices nuevas*.xlsx"

COSTES_PERFILES_PATH = r"\\Dimac2023\datos\DATOS DIMAC\USERS\INTERCAMBIO\RAUL\xavi\PROCES DE COMPRES ADRIA\EXTRUSORES\MATRICES EXTRUSIÓN\COSTES PERFILES NUEVOS\GENERAR COSTES PERFILES NUEVOS.xlsx"
COSTES_PERFILES_SHEET = "COSTE-TARIFA PERFIL (2)"


def localizar_excel_jm(carpeta=JM_CARPETA, patron=JM_PATRON):
    """Devuelve la ruta del Excel de José María más reciente en la carpeta de red."""
    import glob
    import os
    cands = [f for f in glob.glob(os.path.join(carpeta, patron))
             if not os.path.basename(f).startswith("~$")]
    if not cands:
        raise FileNotFoundError(
            f"No encuentro el Excel de José María ({patron}) en:\n  {carpeta}\n"
            "¿Estás en la red de la oficina?")
    cands.sort(key=os.path.getmtime, reverse=True)
    return cands[0]


# Mapeo grupo Geinfor → clave de acabado en Excel costes (BRUTO/BLANCO/etc.)
_GRUPO_FINISH = {
    0: 'BRUTO', 1: 'BLANCO', 2: 'BLANCO', 3: 'BLANCO',
    4: 'ANOD SPC', 5: 'RALSPC', 6: 'RALSTD', 7: 'RALSPC',
    8: 'MAD', 9: 'MAD',
}


def cargar_costes_perfiles(ruta=COSTES_PERFILES_PATH):
    """Lee el Excel de costes de perfiles y devuelve un dict base_code -> datos."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
    except Exception as e:
        print(f"AVISO: No se pudo cargar el Excel de costes de perfiles: {e}")
        return {}
    ws = wb[COSTES_PERFILES_SHEET]

    finish_order = ['BRUTO', 'BLANCO', 'PLATA', 'RALSTD', 'RALSPC', 'ANOD SPC', 'MAD']

    # Fila 21: headers de grupo (COSTE BRUTO, PVP TARIFA). Fila 22: subheaders (BRUTO, BLANCO…)
    row21 = list(ws[21])
    row22 = list(ws[22])
    coste_bruto_col = None
    pvp_tarifa_col = None
    for cell in row21:
        if cell.value and 'COSTE BRUTO' in str(cell.value).upper():
            coste_bruto_col = cell.column - 1
        if cell.value and 'PVP TARIFA' in str(cell.value).upper() and pvp_tarifa_col is None:
            pvp_tarifa_col = cell.column - 1

    cost_finish_col = {}
    pvp_finish_col = {}
    if coste_bruto_col is not None and pvp_tarifa_col is not None:
        for i in range(coste_bruto_col, pvp_tarifa_col):
            v = row22[i].value
            if v and str(v).strip().upper() in [f.upper() for f in finish_order]:
                cost_finish_col[str(v).strip().upper()] = i
        # Solo las 7 columnas exactas del bloque PVP TARIFA (X a AD)
        for i in range(pvp_tarifa_col, pvp_tarifa_col + len(finish_order)):
            if i >= len(row22):
                break
            v = row22[i].value
            if v and str(v).strip().upper() in [f.upper() for f in finish_order]:
                pvp_finish_col[str(v).strip().upper()] = i

    # Mapeo grupo Geinfor → clave finish en Excel
    result = {}
    for row in ws.iter_rows(min_row=23, values_only=True):
        ref = str(row[1]).strip() if row[1] else None  # col B = index 1
        if not ref or ref in ('None', ''):
            continue

        def get(idx):
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            if v is None:
                return None
            try:
                return float(v)
            except Exception:
                return None

        def gets(idx):
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return str(v).strip() if v not in (None, '') else None

        costes = {finish: get(cidx) for finish, cidx in cost_finish_col.items()}
        pvps = {finish: get(cidx) for finish, cidx in pvp_finish_col.items()}

        result[ref] = {
            'peso':          get(2),    # col C
            'perimetro':     get(3),    # col D - PERIMETRO TOTAL
            'perimetro_ext': get(4),    # col E - PERIMETRO EXTERIOR
            'caras':         get(5),    # col F - Nº CARAS
            'poli1':         gets(7),   # col H
            'cant_poli1':    get(8),    # col I
            'poli2':         gets(9),   # col J
            'cant_poli2':    get(10),   # col K
            'comp1':         gets(11),  # col L
            'cant_comp1':    get(12),   # col M
            'comp2':         gets(13),  # col N
            'cant_comp2':    get(14),   # col O
            'mecanizado':    gets(15),  # col P
            'costes': costes,
            'pvps': pvps,
        }

    # Leer el coste de ENSAMBLADO (F2) desde col Q fila 2 de la hoja principal
    try:
        ensamblado_coste = None
        for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
            v = row[16] if len(row) > 16 else None  # col Q = index 16
            try:
                ensamblado_coste = float(v) if v is not None else None
            except Exception:
                pass
        result['__ensamblado_coste__'] = ensamblado_coste
    except Exception:
        pass

    return result


def cargar_tablas_acabados(ruta=COSTES_PERFILES_PATH):
    """Lee las 3 hojas de referencia del Excel de costes y devuelve dicts de costes."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
    except Exception as e:
        print(f"AVISO: No se pudo cargar tablas de acabados: {e}")
        return {}, {}, {}

    # ---- DATOS COSTES ACABADOS: grupo -> coste referencia ----
    coste_acabado = {}
    if 'DATOS COSTES ACABADOS' in wb.sheetnames:
        ws_ac = wb['DATOS COSTES ACABADOS']
        grupos_arts = {}
        for row in ws_ac.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            nombre = str(row[0]).strip()
            grupo = row[9] if len(row) > 9 else None
            coste = row[4] if len(row) > 4 else None
            estado = row[11] if len(row) > 11 else None
            try:
                grupo = int(grupo)
            except Exception:
                continue
            try:
                coste = float(coste)
            except Exception:
                continue
            if estado == 60:
                continue
            grupos_arts.setdefault(grupo, []).append((nombre, coste))

        # Regla por grupo
        for g, arts in grupos_arts.items():
            costes_g = [c for _, c in arts]
            if g == 9:
                coste_acabado[g] = 6.0  # fijo Metalmadera
            elif g == 4:
                inox = next((c for n, c in arts if 'INOX LIJADO REPULIDO' in n.upper()), None)
                coste_acabado[g] = inox if inox is not None else max(costes_g)
            else:
                coste_acabado[g] = max(costes_g)

        # Grupos 5 y 7: promedio de artículos activos con coste >= 4.29 de AMBOS grupos
        arts_57 = [c for g in (5, 7) for _, c in grupos_arts.get(g, []) if c >= 4.29]
        if arts_57:
            prom_57 = round(sum(arts_57) / len(arts_57), 2)
            coste_acabado[5] = prom_57
            coste_acabado[7] = prom_57

    # ---- DATOS COSTES POLIAMIDAS: codigo -> coste_m_l ----
    coste_poli = {}
    if 'DATOS COSTES POLIAMIDAS' in wb.sheetnames:
        ws_po = wb['DATOS COSTES POLIAMIDAS']
        for row in ws_po.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            cod = str(row[0]).strip()
            coste_ml = row[6] if len(row) > 6 else None  # col G = COSTE M/L
            try:
                coste_poli[cod] = float(coste_ml)
            except Exception:
                pass

    # ---- DATOS COSTES MECANIZADOS: codigo -> coste_m_l ----
    coste_mec = {}
    if 'DATOS COSTES MECANIZADOS' in wb.sheetnames:
        ws_me = wb['DATOS COSTES MECANIZADOS']
        for row in ws_me.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            cod = str(row[0]).strip()
            coste_ml = row[5] if len(row) > 5 else None  # col F = COSTE M/L
            try:
                coste_mec[cod] = float(coste_ml)
            except Exception:
                pass

    return coste_acabado, coste_poli, coste_mec


_MCOLS = ("CODIGO_ARTICULO,DENOMINACION,DENOMINACION_2,CLASE,TIPO,NORMA,COEFICIENTE,"
          "UNIDAD_INTERNA,UNIDAD_EXTERNA,FAMILIAPRODUCTO,GRUPO,MEDIDA,UNIDADES_ARTICULO,"
          "ESTADO,FECHACREACION,FECHAMODIFICACION,"
          "PARTIDAARANCELARIA,CONVERSION_M_KG,"
          "GESTIONARVARIABLES,ESSUBCONTRATACIO,ESKIT,NOGESTIONASTOCK,CALCULOCANTFORMULA,"
          "TIPOARTMODULAR,ARTICULOBASE")


def _leer_in(cn, tabla, clave, codigos, cols, extra=""):
    """SELECT cols FROM tabla WHERE clave IN (codigos) [extra], troceado para no pasarse
    del límite de parámetros. Devuelve un único DataFrame."""
    partes = []
    cod = list(codigos)
    for i in range(0, len(cod), 500):
        trozo = cod[i:i + 500]
        ph = ",".join("?" * len(trozo))
        sql = f"SELECT {cols} FROM DB2ADMIN.{tabla} WHERE {clave} IN ({ph}) {extra}"
        partes.append(pd.read_sql(sql, cn, params=trozo))
    if partes:
        return pd.concat(partes, ignore_index=True)
    return pd.DataFrame(columns=[c.strip() for c in cols.split(",")])


def _conteo_plazos_db2(cn, proveedores):
    """Conteo de PLAZOENTREGA por proveedor (PRINCIPAL=1) para los proveedores dados.
    Devuelve df PROVEEDOR / PLAZOENTREGA / N (vía GROUP BY, ligero)."""
    partes = []
    provs = list(proveedores)
    for i in range(0, len(provs), 500):
        t = provs[i:i + 500]
        ph = ",".join("?" * len(t))
        sql = (f"SELECT PROVEEDOR, PLAZOENTREGA, COUNT(*) AS N FROM DB2ADMIN.ARTICULOS_PROVEEDOR "
               f"WHERE PRINCIPAL=1 AND PROVEEDOR IN ({ph}) GROUP BY PROVEEDOR, PLAZOENTREGA")
        partes.append(pd.read_sql(sql, cn, params=t))
    if partes:
        return pd.concat(partes, ignore_index=True)
    return pd.DataFrame(columns=["PROVEEDOR", "PLAZOENTREGA", "N"])


def plazo_habitual_por_proveedor(cnt, min_muestras=5, dominancia=0.6):
    """A partir del conteo (PROVEEDOR/PLAZOENTREGA/N), devuelve {proveedor: plazo_habitual}
    solo para proveedores con suficientes artículos y un plazo claramente dominante."""
    out = {}
    if cnt is None or len(cnt) == 0:
        return out
    c = cnt.copy()
    c["PROVEEDOR"] = pd.to_numeric(c["PROVEEDOR"], errors="coerce")
    c["PLAZOENTREGA"] = pd.to_numeric(c["PLAZOENTREGA"], errors="coerce")
    c["N"] = pd.to_numeric(c["N"], errors="coerce").fillna(0)
    c = c.dropna(subset=["PROVEEDOR", "PLAZOENTREGA"])
    for pid, g in c.groupby("PROVEEDOR"):
        total = g["N"].sum()
        if total >= min_muestras:
            top = g.sort_values("N", ascending=False).iloc[0]
            if top["N"] / total >= dominancia:
                out[int(pid)] = top["PLAZOENTREGA"]
    return out


def leer_db2(desde=None, hasta=None, articulo=None, dsn=DSN_DB2):
    """Lee de Geinfor (DB2, SOLO LECTURA) solo lo necesario: el universo de accesorios
    (familia != 1,2) + catálogos (tipo 91), y de las tablas relacionadas únicamente las
    filas de los artículos EN ALCANCE y con los filtros que el validador ya aplica
    (proveedor PRINCIPAL=1, tarifa 40, característica 0). Devuelve la misma estructura
    que las hojas del Excel."""
    import pyodbc
    import warnings
    warnings.filterwarnings("ignore", message="pandas only supports SQLAlchemy")
    cn = pyodbc.connect("DSN=" + dsn, timeout=30, readonly=True)  # readonly: regla de oro
    try:
        # Maestro: en modo un-artículo, ese código (sea cual sea su familia) + catálogos,
        # para poder informar si resulta ser un perfil. En lote, accesorios + catálogos.
        if articulo is not None:
            m = pd.read_sql(
                f"SELECT {_MCOLS} FROM DB2ADMIN.MAESTRO_DE_ARTICULOS "
                "WHERE TIPO=91 OR (TIPO IN (1,92) AND FAMILIAPRODUCTO=105) OR CODIGO_ARTICULO=?",
                cn, params=[articulo])
        else:
            m = pd.read_sql(
                f"SELECT {_MCOLS} FROM DB2ADMIN.MAESTRO_DE_ARTICULOS "
                "WHERE TIPO=91 OR ESTADO<>60", cn)
        # Códigos EN ALCANCE: un solo artículo (--articulo) o los accesorios del rango.
        if articulo is not None:
            codigos = [articulo]
            codigos_p = [articulo]
        else:
            fam = pd.to_numeric(m["FAMILIAPRODUCTO"], errors="coerce")
            fc = pd.to_datetime(m["FECHACREACION"], errors="coerce")
            mask = ~fam.isin([1, 2])
            if desde is not None:
                mask &= fc >= desde
            if hasta is not None:
                mask &= fc <= hasta
            codigos = m.loc[mask, "CODIGO_ARTICULO"].astype(str).str.strip().tolist()

            fam_p = pd.to_numeric(m["FAMILIAPRODUCTO"], errors="coerce")
            mask_p = fam_p.isin([1, 2])
            if desde is not None:
                fc_p = pd.to_datetime(m["FECHACREACION"], errors="coerce")
                mask_p &= fc_p >= desde
            if hasta is not None:
                fc_p = pd.to_datetime(m["FECHACREACION"], errors="coerce")
                mask_p &= fc_p <= hasta
            codigos_p = m.loc[mask_p, "CODIGO_ARTICULO"].astype(str).str.strip().tolist()

        tar = _leer_in(cn, "ARTICULOS_TARIFABLES", "ARTICULO", codigos,
                       "ARTICULO,TARIFA,PRECIO_VENTA_PUBLICO,UNIDADVENTA,COEFICIENTE,IVA", "AND TARIFA=40")
        prov = _leer_in(cn, "ARTICULOS_PROVEEDOR", "ARTICULO", codigos,
                        "ARTICULO,PROVEEDOR,PRINCIPAL,CODIGO_PROVEEDOR,COSTEBRUTO,COSTEINTERNO,PLAZOENTREGA,"
                        "LOTEHABITUAL,UNIDADEXTERNA,COEFICIENTE,PIVA,FORMULA", "AND PRINCIPAL=1")
        alm = _leer_in(cn, "ALMACEN_ARTICULOS", "ARTICULO", codigos, "ARTICULO,STOCKMINIMO,LOTEFABRICACION")
        artvar = _leer_in(cn, "ARTICULOVARIABLE", "ARTICULO", codigos, "ARTICULO,VARIABLE")
        artasoc = _leer_in(cn, "ARTICULOSASOCIADOS", "ARTICULO", codigos,
                           "ARTICULO,ARTICULOASOCIADO,CANTIDAD,INTEGRACIONAUTO")
        car = _leer_in(cn, "TCARAC_ART", "ARTICULO", codigos,
                       "ARTICULO,CARACTERISTICA,VALOR", "AND CARACTERISTICA=0")
        est = _leer_in(cn, "MAESTRO_ESTRUCTURAS", "ARTSUPERIOR", codigos,
                       "ARTSUPERIOR,ARTCOMPONENTE,CANTIDAD,FORMULA,CANTIDADFIJA,RECALCULARFORMULA")
        # Perfiles (F1/F2): tablas relacionadas
        tar_p = _leer_in(cn, "ARTICULOS_TARIFABLES", "ARTICULO", codigos_p,
                         "ARTICULO,TARIFA,PRECIO_VENTA_PUBLICO,UNIDADVENTA,COEFICIENTE,IVA", "AND TARIFA=40")
        prov_p = _leer_in(cn, "ARTICULOS_PROVEEDOR", "ARTICULO", codigos_p,
                          "ARTICULO,PROVEEDOR,PRINCIPAL,CODIGO_PROVEEDOR,COSTEBRUTO,COSTEINTERNO,PLAZOENTREGA,"
                          "LOTEHABITUAL,UNIDADEXTERNA,COEFICIENTE,PIVA,FORMULA", "AND PRINCIPAL=1")
        est_p = _leer_in(cn, "MAESTRO_ESTRUCTURAS", "ARTSUPERIOR", codigos_p,
                         "ARTSUPERIOR,ARTCOMPONENTE,CANTIDAD,FORMULA,CANTIDADFIJA,RECALCULARFORMULA")
        alm_p = _leer_in(cn, "ALMACEN_ARTICULOS", "ARTICULO", codigos_p,
                         "ARTICULO,ALMACEN,STOCKMINIMO,STOCKMAXIMO,LOTEFABRICACION")
        artvar_p = _leer_in(cn, "ARTICULOVARIABLE", "ARTICULO", codigos_p, "ARTICULO,VARIABLE")
        car_p = _leer_in(cn, "TCARAC_ART", "ARTICULO", codigos_p,
                         "ARTICULO,CARACTERISTICA,VALOR")
        clases = pd.read_sql("SELECT CODIGO, DENOMINACION FROM DB2ADMIN.TABLA_CLASE_ARTICULO", cn)
        mprov = pd.read_sql("SELECT CODIGO_PROVEEDOR, TIPO_CTA_COMPRA FROM DB2ADMIN.MAESTRO_PROVEEDORES", cn)
        # Conteo de plazos por proveedor (para comparar contra el plazo habitual del proveedor)
        provs = pd.to_numeric(prov["PROVEEDOR"], errors="coerce").dropna().astype(int).unique().tolist()
        cnt_plazos = _conteo_plazos_db2(cn, provs)
    finally:
        cn.close()
    return m, tar, prov, alm, car, est, clases, cnt_plazos, mprov, artvar, artasoc, tar_p, prov_p, est_p, alm_p, artvar_p, car_p


_ESTILO_HTML = """<style>
  :root{--err:#c0392b;--errbg:#fdecea;--warn:#b9770e;--warnbg:#fdf3e3;--ok:#1e8449;--okbg:#eafaf1;}
  *{box-sizing:border-box} body{font-family:Segoe UI,system-ui,Arial,sans-serif;margin:0;background:#f4f6f8;color:#222}
  .wrap{max-width:820px;margin:0 auto;padding:24px}
  .head{background:#fff;border-radius:14px;padding:22px 24px;box-shadow:0 1px 4px rgba(0,0,0,.08)}
  .cod{font-size:30px;font-weight:700;letter-spacing:.5px}
  .denom{font-size:18px;color:#444;margin-top:2px}
  .meta{color:#777;font-size:13px;margin-top:10px}
  .banner{margin:18px 0;padding:14px 18px;border-radius:12px;font-size:18px;font-weight:600}
  .banner.ok{background:var(--okbg);color:var(--ok)}
  .banner.bad{background:#fff;box-shadow:0 1px 4px rgba(0,0,0,.06)}
  .pill{display:inline-block;padding:5px 12px;border-radius:20px;font-size:14px;margin-right:8px;color:#fff}
  .pill.err{background:var(--err)} .pill.warn{background:var(--warn)} .pill.ok{background:var(--ok)}
  h2.th{font-size:15px;text-transform:uppercase;letter-spacing:.5px;margin:22px 0 10px}
  h2.th.err{color:var(--err)} h2.th.warn{color:var(--warn)}
  .card{background:#fff;border-left:5px solid #ccc;border-radius:10px;padding:12px 16px;margin-bottom:10px;box-shadow:0 1px 3px rgba(0,0,0,.06)}
  .card.err{border-color:var(--err)} .card.warn{border-color:var(--warn)} .card.ok{border-color:var(--ok);color:var(--ok)}
  .art.ok{border-left:5px solid var(--ok);opacity:.85}
  .campo{font-weight:700;font-size:15px} .detalle{color:#555;margin-top:3px;font-size:14px}
  .foot{color:#999;font-size:12px;margin-top:26px;text-align:center}
  .art{background:#fff;border-radius:12px;padding:16px 20px;margin:14px 0;box-shadow:0 1px 4px rgba(0,0,0,.07)}
  .acod{font-size:20px;font-weight:700} .adenom{color:#555;font-size:14px;margin-bottom:8px}
</style>"""


def informe_mensaje(cod, titulo, mensaje, ruta, clase="bad"):
    """Informe HTML simple con un único mensaje (p.ej. 'es un perfil' o 'no encontrado')."""
    e = _html.escape
    ahora = pd.Timestamp.now().strftime("%d/%m/%Y %H:%M")
    doc = f"""<!doctype html><html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>Validación {e(cod)}</title>
{_ESTILO_HTML}</head><body><div class="wrap">
  <div class="head"><div class="cod">{e(cod)}</div></div>
  <div class="banner {clase}" style="font-size:17px">{e(titulo)}</div>
  <div class="card"><div class="detalle" style="font-size:15px">{e(mensaje)}</div></div>
  <div class="foot">Validador de accesorios Dimac · datos en directo de Geinfor (solo lectura) · {ahora}</div>
</div></body></html>"""
    with open(ruta, "w", encoding="utf-8") as f:
        f.write(doc)


def generar_informe_lote_html(sub, out, etiqueta, ruta):
    """Informe visual (HTML) de la validación de un LOTE de altas: un bloque por artículo
    con incidencias (errores y avisos), ordenados los que tienen errores primero."""
    def esc(x):
        return _html.escape(str(x))
    total = len(sub)
    n_err = int((out["NIVEL"] == "ERROR").sum()) if len(out) else 0
    n_av = int((out["NIVEL"] == "AVISO").sum()) if len(out) else 0
    con_inc = out["CODIGO"].nunique() if len(out) else 0
    correctos = total - con_inc
    info = sub.set_index("CODIGO_ARTICULO")

    codigos_inc = list(dict.fromkeys(out["CODIGO"].tolist())) if len(out) else []
    def nerr(c):
        return int(((out["CODIGO"] == c) & (out["NIVEL"] == "ERROR")).sum())
    codigos_inc.sort(key=lambda c: (0 if nerr(c) > 0 else 1, c))

    codigos_ok = [c for c in sub["CODIGO_ARTICULO"].tolist() if c not in set(codigos_inc)]

    bloques = []
    for c in codigos_inc:
        filas = out[out["CODIGO"] == c]
        denom = esc(info.loc[c, "DENOMINACION"]) if c in info.index else ""
        cards = []
        for _, r in filas.iterrows():
            clase = "err" if r["NIVEL"] == "ERROR" else "warn"
            nombre = NOMBRES_CAMPO.get(r["CAMPO"], r["CAMPO"])
            cards.append(f'<div class="card {clase}"><div class="campo">{esc(nombre)}</div>'
                         f'<div class="detalle">{esc(r["DETALLE"])}</div></div>')
        bloques.append(f'<div class="art"><div class="acod">{esc(c)}</div>'
                       f'<div class="adenom">{denom}</div>{"".join(cards)}</div>')
    for c in codigos_ok:
        denom = esc(info.loc[c, "DENOMINACION"]) if c in info.index else ""
        bloques.append(f'<div class="art ok"><div class="acod">{esc(c)}</div>'
                       f'<div class="adenom">{denom}</div>'
                       f'<div class="card ok"><div class="detalle">✓ Sin incidencias</div></div></div>')

    ahora = pd.Timestamp.now().strftime("%d/%m/%Y %H:%M")
    pills = (f'<span class="pill ok">{correctos} correctos</span>'
             f'<span class="pill err">{n_err} errores</span>'
             f'<span class="pill warn">{n_av} avisos</span>')
    cuerpo = "".join(bloques) if bloques else '<div class="banner ok">✓ Sin incidencias</div>'
    doc = f"""<!doctype html><html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>Validación de altas</title>
{_ESTILO_HTML}</head><body><div class="wrap">
  <div class="head"><div class="cod">Validación de altas</div>
    <div class="denom">{esc(etiqueta)}</div>
    <div class="meta">{total} accesorios evaluados</div></div>
  <div class="banner bad">{pills}</div>
  {cuerpo}
  <div class="foot">Validador de accesorios Dimac · datos en directo de Geinfor (solo lectura) · {ahora}</div>
</div></body></html>"""
    with open(ruta, "w", encoding="utf-8") as f:
        f.write(doc)


def generar_informe_html(cod, fila, inc_art, ruta):
    """Genera un informe visual (HTML) de la validación de UN artículo."""
    def esc(x):
        return _html.escape(str(x))
    errores = [r for r in inc_art if r[2] == "ERROR"]
    avisos = [r for r in inc_art if r[2] == "AVISO"]
    denom = esc(fila.get("DENOMINACION", ""))
    cab = (f"Familia {esc(fila.get('FAMILIAPRODUCTO'))} · Tipo {esc(fila.get('TIPO'))} · "
           f"Grupo {esc(fila.get('GRUPO'))} · Estado {esc(fila.get('ESTADO'))} · "
           f"Clase {esc(fila.get('CLASE'))}")
    fcre = fila.get("FECHACREACION")
    fcre = esc(pd.to_datetime(fcre).date()) if pd.notna(fcre) else "-"

    if not errores and not avisos:
        banner = '<div class="banner ok">✓ Artículo correcto — sin incidencias</div>'
    else:
        partes = []
        if errores:
            partes.append(f'<span class="pill err">{len(errores)} error(es)</span>')
        if avisos:
            partes.append(f'<span class="pill warn">{len(avisos)} aviso(s)</span>')
        banner = f'<div class="banner bad">{" ".join(partes)}</div>'

    def tarjetas(lista, clase):
        out = []
        for _, campo, _niv, detalle in lista:
            nombre = NOMBRES_CAMPO.get(campo, campo)
            out.append(f'<div class="card {clase}"><div class="campo">{esc(nombre)}</div>'
                       f'<div class="detalle">{esc(detalle)}</div></div>')
        return "\n".join(out)

    secciones = ""
    if errores:
        secciones += f'<h2 class="th err">✕ Errores ({len(errores)})</h2>{tarjetas(errores, "err")}'
    if avisos:
        secciones += f'<h2 class="th warn">⚠ Avisos ({len(avisos)})</h2>{tarjetas(avisos, "warn")}'

    ahora = pd.Timestamp.now().strftime("%d/%m/%Y %H:%M")
    doc = f"""<!doctype html><html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Validación {esc(cod)}</title>
{_ESTILO_HTML}</head><body><div class="wrap">
  <div class="head">
    <div class="cod">{esc(cod)}</div>
    <div class="denom">{denom}</div>
    <div class="meta">{cab} · Creado: {fcre}</div>
  </div>
  {banner}
  {secciones}
  <div class="foot">Validador de accesorios Dimac · datos en directo de Geinfor (solo lectura) · {ahora}</div>
</div></body></html>"""
    with open(ruta, "w", encoding="utf-8") as f:
        f.write(doc)


def main():
    # Modos:
    #   Excel:        python validador.py REPASO.xlsx JM.xlsx [DESDE] [HASTA]
    #   DB2 (rango):  python validador.py --db JM.xlsx [DESDE] [HASTA]
    #   Un artículo:  python validador.py --db JM.xlsx --articulo CODIGO
    args = sys.argv[1:]
    articulo = None
    if "--articulo" in args:
        i = args.index("--articulo")
        articulo = args[i + 1].strip()
        del args[i:i + 2]

    if args[0] == "--db":
        # El Excel de JM se localiza solo en la red; opcionalmente se puede pasar uno explícito.
        rest = args[1:]
        if rest and rest[0].lower().endswith(".xlsx"):
            jmpath = rest[0]
            rest = rest[1:]
        else:
            jmpath = localizar_excel_jm()
            print(f"Excel de José María (red): {jmpath}")
        desde = pd.to_datetime(rest[0]) if len(rest) > 0 else None
        hasta = pd.to_datetime(rest[1]) if len(rest) > 1 else None
        print(f"Leyendo Geinfor en directo (DB2, DSN {DSN_DB2})...")
        m, tar, prov, alm, car, est, clases, cnt_plazos, mprov, artvar, artasoc, tar_p, prov_p, est_p, alm_p, artvar_p, car_p = leer_db2(desde, hasta, articulo)
    else:
        repaso, jmpath = args[0], args[1]
        desde = pd.to_datetime(args[2]) if len(args) > 2 else None
        hasta = pd.to_datetime(args[3]) if len(args) > 3 else None
        xl = pd.ExcelFile(repaso)
        m = pd.read_excel(xl, "MAESTRO DE ARTICULOS", dtype={"CODIGO_ARTICULO": str})
        tar = pd.read_excel(xl, "ARTICULOS TARIFABLES", dtype={"ARTICULO": str})
        prov = pd.read_excel(xl, "ARTICULOS PROVEEDOR", dtype={"ARTICULO": str})
        alm = pd.read_excel(xl, "ALMACEN ARTICULOS", dtype={"ARTICULO": str})
        car = pd.read_excel(xl, "TABLA CARACTERISTICAS", dtype=str)
        est = pd.read_excel(xl, "MAESTRO DE ESTRUCTURAS", dtype=str)
        clases = pd.read_excel(xl, "CLASE ARTICULO", dtype=str)
        # Conteo de plazos por proveedor desde la tabla completa de proveedores
        _pp = prov[prov["PRINCIPAL"].astype(str) == "1"]
        cnt_plazos = _pp.groupby(["PROVEEDOR", "PLAZOENTREGA"]).size().reset_index(name="N")
        mprov = pd.DataFrame(columns=["CODIGO_PROVEEDOR", "TIPO_CTA_COMPRA"])  # no disponible en modo Excel
        artvar = pd.DataFrame(columns=["ARTICULO", "VARIABLE"])
        artasoc = pd.DataFrame(columns=["ARTICULO", "ARTICULOASOCIADO", "CANTIDAD", "INTEGRACIONAUTO"])
        tar_p = pd.DataFrame(columns=["ARTICULO", "TARIFA", "PRECIO_VENTA_PUBLICO", "UNIDADVENTA", "COEFICIENTE", "IVA"])
        prov_p = pd.DataFrame(columns=["ARTICULO", "PROVEEDOR", "PRINCIPAL", "CODIGO_PROVEEDOR", "COSTEBRUTO",
                                        "COSTEINTERNO", "PLAZOENTREGA", "LOTEHABITUAL", "UNIDADEXTERNA",
                                        "COEFICIENTE", "PIVA", "FORMULA"])
        est_p = pd.DataFrame(columns=["ARTSUPERIOR", "ARTCOMPONENTE", "CANTIDAD", "FORMULA",
                                       "CANTIDADFIJA", "RECALCULARFORMULA"])
        alm_p = pd.DataFrame(columns=["ARTICULO", "ALMACEN", "STOCKMINIMO", "STOCKMAXIMO", "LOTEFABRICACION"])
        artvar_p = pd.DataFrame(columns=["ARTICULO", "VARIABLE"])
        car_p = pd.DataFrame(columns=["ARTICULO", "CARACTERISTICA", "VALOR"])

    jm = pd.read_excel(jmpath, sheet_name="RELLENAR JOSE MARIA", dtype=str)
    jm["ARTÍCULO"] = jm["ARTÍCULO"].astype(str).str.strip()
    en_jm = set(jm["ARTÍCULO"])

    m["CODIGO_ARTICULO"] = m["CODIGO_ARTICULO"].astype(str).str.strip()
    m["FECHACREACION"] = pd.to_datetime(m["FECHACREACION"], errors="coerce")
    m["FECHAMODIFICACION"] = pd.to_datetime(m["FECHAMODIFICACION"], errors="coerce")

    # Universo ACCESORIOS: familia distinta de 1, 2 y 105.
    # Familia 105 = chapas y paneles — tienen reglas propias, se validarán aparte.
    # Estado 60 = artículo de baja — no se valida.
    fam = m["FAMILIAPRODUCTO"].apply(ni)
    estado_col = m["ESTADO"].apply(ni)
    es_accesorio = ~fam.isin([1, 2, 105])
    es_activo = estado_col != 60
    if articulo is not None:
        # Un artículo: se valida si es accesorio O si JM lo lista (lo trata como accesorio,
        # aunque su familia esté mal puesta a 1/2; así la regla FAMILIA detecta el error).
        # Si está de baja (estado 60) se avisa pero no se validan sus campos.
        sub = m[(m["CODIGO_ARTICULO"] == articulo) & es_accesorio &
                (es_accesorio | m["CODIGO_ARTICULO"].isin(en_jm))].copy()
        if not sub.empty and ni(sub.iloc[0].get("ESTADO")) == 60:
            print(f"{articulo}: artículo de baja (estado 60) — no se valida.")
            return pd.DataFrame(columns=["CODIGO", "CAMPO", "NIVEL", "DESCRIPCION"])
    else:
        # Solo artículos CREADOS (altas nuevas), filtrando por FECHACREACION.
        sub = m[es_accesorio & es_activo].copy()
        if desde is not None:
            sub = sub[sub["FECHACREACION"] >= desde]
        if hasta is not None:
            sub = sub[sub["FECHACREACION"] <= hasta]

    # Familia 105 (chapas y paneles): subconjunto con reglas propias
    if articulo is not None:
        sub_105 = m[(m["CODIGO_ARTICULO"] == articulo) & fam.isin([105])].copy()
        if not sub_105.empty and ni(sub_105.iloc[0].get("ESTADO")) == 60:
            print(f"{articulo}: artículo de baja (estado 60) — no se valida.")
            return pd.DataFrame(columns=["CODIGO", "CAMPO", "NIVEL", "DESCRIPCION"])
    else:
        sub_105 = m[fam.isin([105]) & es_activo].copy()
        if desde is not None:
            sub_105 = sub_105[sub_105["FECHACREACION"] >= desde]
        if hasta is not None:
            sub_105 = sub_105[sub_105["FECHACREACION"] <= hasta]

    # Familia 1 y 2 (perfiles fríos y rotura térmica)
    if articulo is not None:
        sub_p = m[(m["CODIGO_ARTICULO"] == articulo) & fam.isin([1, 2])].copy()
    else:
        sub_p = m[fam.isin([1, 2]) & es_activo].copy()
        if desde is not None:
            sub_p = sub_p[sub_p["FECHACREACION"] >= desde]
        if hasta is not None:
            sub_p = sub_p[sub_p["FECHACREACION"] <= hasta]

    # --- índices de apoyo (siempre proveedor PRINCIPAL = 1) ---
    prov["ARTICULO"] = prov["ARTICULO"].astype(str).str.strip()
    pr = prov[prov["PRINCIPAL"].astype(str) == "1"].set_index("ARTICULO")
    # Coste para el margen = COSTEINTERNO (coste ACTUAL neto: bruto con dto y recargo aplicados).
    # NO usar COSTEBRUTO (es el precio de tarifa del proveedor antes de descuento).
    coste_pr = pd.to_numeric(pr["COSTEINTERNO"], errors="coerce").to_dict()
    costebruto_pr = pd.to_numeric(pr["COSTEBRUTO"], errors="coerce").to_dict()
    ref_prov_pr = pr["CODIGO_PROVEEDOR"].astype(str).str.strip().to_dict() if "CODIGO_PROVEEDOR" in pr.columns else {}
    plazo_pr = pd.to_numeric(pr["PLAZOENTREGA"], errors="coerce").to_dict()
    lote_pr = pd.to_numeric(pr["LOTEHABITUAL"], errors="coerce").to_dict()
    uext_pr = pd.to_numeric(pr["UNIDADEXTERNA"], errors="coerce").to_dict()
    proveedor_pr = pr["PROVEEDOR"].astype(str).to_dict()  # cod_articulo -> cod_proveedor
    piva_pr = pd.to_numeric(pr["PIVA"], errors="coerce").to_dict()
    formula_pr = pd.to_numeric(pr["FORMULA"], errors="coerce").to_dict()
    con_principal = set(pr.index)

    plazo_habitual = plazo_habitual_por_proveedor(cnt_plazos)

    # TIPO_CTA_COMPRA por proveedor: 0=nacional(ES), 1=intracomunitario, 2=extracomunitario
    mprov["CODIGO_PROVEEDOR"] = pd.to_numeric(mprov["CODIGO_PROVEEDOR"], errors="coerce")
    tipo_cta_prov = mprov.set_index("CODIGO_PROVEEDOR")["TIPO_CTA_COMPRA"].to_dict()

    tar["ARTICULO"] = tar["ARTICULO"].astype(str).str.strip()
    t40 = tar[pd.to_numeric(tar["TARIFA"], errors="coerce") == TARIFA_PVP]
    pvp40 = pd.to_numeric(t40.set_index("ARTICULO")["PRECIO_VENTA_PUBLICO"], errors="coerce").to_dict()
    uventa_tar = pd.to_numeric(t40.set_index("ARTICULO")["UNIDADVENTA"], errors="coerce").to_dict()
    iva_tar40 = pd.to_numeric(t40.set_index("ARTICULO")["IVA"], errors="coerce").to_dict()
    coef_tar40 = pd.to_numeric(t40.set_index("ARTICULO")["COEFICIENTE"], errors="coerce").to_dict() if "COEFICIENTE" in t40.columns else {}

    car["ARTICULO"] = car["ARTICULO"].astype(str).str.strip()
    car0 = set(car[(car["CARACTERISTICA"].astype(str) == "0") &
                   (car["VALOR"].astype(str).str.strip().replace("nan", "") != "")]["ARTICULO"])

    alm["ARTICULO"] = alm["ARTICULO"].astype(str).str.strip()
    stockmin = pd.to_numeric(alm.set_index("ARTICULO")["STOCKMINIMO"], errors="coerce").to_dict()
    lote_alm = pd.to_numeric(alm.set_index("ARTICULO")["LOTEFABRICACION"], errors="coerce").to_dict() if "LOTEFABRICACION" in alm.columns else {}

    # Artículo variable: set de artículos que tienen VARIABLE=10
    artvar["ARTICULO"] = artvar["ARTICULO"].astype(str).str.strip()
    artvar_10 = set(artvar[pd.to_numeric(artvar["VARIABLE"], errors="coerce") == 10]["ARTICULO"])

    # Artículos asociados: dict cod -> set de articulosasociados
    artasoc["ARTICULO"] = artasoc["ARTICULO"].astype(str).str.strip()
    artasoc["ARTICULOASOCIADO"] = artasoc["ARTICULOASOCIADO"].astype(str).str.strip()
    asoc_por_art = artasoc.groupby("ARTICULO")["ARTICULOASOCIADO"].apply(set).to_dict()
    # Dict cod -> lista de dicts con los datos de cada línea de asociado
    asoc_lineas = {cod: grupo[["ARTICULOASOCIADO", "CANTIDAD", "INTEGRACIONAUTO"]].to_dict("records")
                   for cod, grupo in artasoc.groupby("ARTICULO")
                   if all(c in artasoc.columns for c in ["CANTIDAD", "INTEGRACIONAUTO"])}

    # Estructura: componentes por artículo padre
    est["ARTSUPERIOR"] = est["ARTSUPERIOR"].astype(str).str.strip()
    est["ARTCOMPONENTE"] = est["ARTCOMPONENTE"].astype(str).str.strip()
    componentes = est.groupby("ARTSUPERIOR")["ARTCOMPONENTE"].apply(set).to_dict()

    # Catálogos = artículos con TIPO 91 (CATALOGOINTERIORES, MADERA, RALESTANDAR, etc.).
    # En accesorios, el único catálogo válido como componente es CATALOGOINTERIORES;
    # el resto de catálogos son de perfiles.
    tipo_de = pd.to_numeric(m["TIPO"], errors="coerce")
    catalogos = set(m.loc[tipo_de == 91, "CODIGO_ARTICULO"])

    # Apoyo para validación familia 105:
    # grupo de cada catálogo tipo 91 (para validar tipo 92)
    _f105_grupo_de_cat = {
        row["CODIGO_ARTICULO"]: ni(row.get("GRUPO"))
        for _, row in m[tipo_de == 91].iterrows()
    }
    # grupo de cada artículo tipo 92 de familia 105 (para validar tipo 60)
    _f105_grupo_de_92 = {
        row["CODIGO_ARTICULO"]: ni(row.get("GRUPO"))
        for _, row in m[fam.isin([105]) & (tipo_de == 92)].iterrows()
    }
    # Códigos de artículos tipo 1 de familia 105 (chapas base en bruto)
    _f105_tipo1 = set(m.loc[fam.isin([105]) & (tipo_de == 1), "CODIGO_ARTICULO"])

    jm["ARTÍCULO"] = jm["ARTÍCULO"].astype(str).str.strip()
    jm_idx = jm.set_index("ARTÍCULO")
    divis = pd.to_numeric(jm_idx["DIVISOR ACCESORIO"], errors="coerce").to_dict()
    minimo_jm = pd.to_numeric(jm_idx["MINIMO ACCESORIO"], errors="coerce").to_dict()
    estado_jm = pd.to_numeric(jm_idx["ESTADO"], errors="coerce").to_dict()
    familia_jm = pd.to_numeric(jm_idx["FAMILIA"], errors="coerce").to_dict()
    clase_serie_jm = jm_idx["CLASE(SERIE)"].fillna("").astype(str).to_dict()
    partida_jm = pd.to_numeric(jm_idx["PARTIDA ARANCELARIA"], errors="coerce").to_dict()
    proveedor_jm = jm_idx["PROVEEDOR"].fillna("").astype(str).to_dict() if "PROVEEDOR" in jm_idx.columns else {}
    en_jm = set(jm["ARTÍCULO"])

    costes_p = cargar_costes_perfiles()
    ensamblado_coste_ref = costes_p.pop('__ensamblado_coste__', None) or 0.60
    coste_acabado, coste_poli, coste_mec = cargar_tablas_acabados()

    # Columnas JM para perfiles: almacén mínimos por grupo
    _jm_min_bruto  = pd.to_numeric(jm_idx.get("MINIMO BRUTO",  pd.Series(dtype=str)), errors="coerce").to_dict()
    _jm_min_blanco = pd.to_numeric(jm_idx.get("MINIMO BLANCO", pd.Series(dtype=str)), errors="coerce").to_dict()
    _jm_min_plata  = pd.to_numeric(jm_idx.get("MINIMO PLATA",  pd.Series(dtype=str)), errors="coerce").to_dict()
    _jm_lote       = pd.to_numeric(jm_idx.get("LOTE HABITUAL", pd.Series(dtype=str)), errors="coerce").to_dict()

    # Índices perfiles: proveedor principal
    prov_p["ARTICULO"] = prov_p["ARTICULO"].astype(str).str.strip()
    pr_p = prov_p[prov_p["PRINCIPAL"].astype(str) == "1"].set_index("ARTICULO")
    costebruto_pr_p  = pd.to_numeric(pr_p["COSTEBRUTO"],    errors="coerce").to_dict()
    costeinterno_pr_p = pd.to_numeric(pr_p["COSTEINTERNO"], errors="coerce").to_dict()
    plazo_pr_p       = pd.to_numeric(pr_p["PLAZOENTREGA"],  errors="coerce").to_dict()
    lote_pr_p        = pd.to_numeric(pr_p["LOTEHABITUAL"],  errors="coerce").to_dict()
    uext_pr_p        = pd.to_numeric(pr_p["UNIDADEXTERNA"], errors="coerce").to_dict()
    proveedor_pr_p   = pr_p["PROVEEDOR"].astype(str).to_dict()
    piva_pr_p        = pd.to_numeric(pr_p["PIVA"],          errors="coerce").to_dict()
    ref_prov_pr_p    = pr_p["CODIGO_PROVEEDOR"].astype(str).str.strip().to_dict() if "CODIGO_PROVEEDOR" in pr_p.columns else {}
    coef_pr_p        = pd.to_numeric(pr_p["COEFICIENTE"],   errors="coerce").to_dict()
    con_principal_p  = set(pr_p.index)

    # Índices perfiles: tarifa 40
    tar_p["ARTICULO"] = tar_p["ARTICULO"].astype(str).str.strip()
    t40_p = tar_p[pd.to_numeric(tar_p["TARIFA"], errors="coerce") == TARIFA_PVP]
    pvp40_p     = pd.to_numeric(t40_p.set_index("ARTICULO")["PRECIO_VENTA_PUBLICO"], errors="coerce").to_dict()
    uventa_p    = pd.to_numeric(t40_p.set_index("ARTICULO")["UNIDADVENTA"],          errors="coerce").to_dict()
    iva_tar40_p = pd.to_numeric(t40_p.set_index("ARTICULO")["IVA"],                  errors="coerce").to_dict()
    coef_tar_p  = pd.to_numeric(t40_p.set_index("ARTICULO")["COEFICIENTE"],          errors="coerce").to_dict()

    # Índices perfiles: estructura
    est_p["ARTSUPERIOR"]   = est_p["ARTSUPERIOR"].astype(str).str.strip()
    est_p["ARTCOMPONENTE"] = est_p["ARTCOMPONENTE"].astype(str).str.strip()
    comp_p_det = est_p.groupby("ARTSUPERIOR").apply(
        lambda g: {row["ARTCOMPONENTE"]: row.to_dict() for _, row in g.iterrows()}
    ).to_dict()

    # Índices perfiles: almacén (almacén 0 = principal)
    alm_p["ARTICULO"] = alm_p["ARTICULO"].astype(str).str.strip()
    alm_p0 = alm_p[pd.to_numeric(alm_p.get("ALMACEN", pd.Series(dtype=str)), errors="coerce") == 0]
    stockmin_p  = pd.to_numeric(alm_p0.set_index("ARTICULO")["STOCKMINIMO"],    errors="coerce").to_dict()
    stockmax_p  = pd.to_numeric(alm_p0.set_index("ARTICULO")["STOCKMAXIMO"],    errors="coerce").to_dict() if "STOCKMAXIMO" in alm_p0.columns else {}
    lote_alm_p  = pd.to_numeric(alm_p0.set_index("ARTICULO")["LOTEFABRICACION"],errors="coerce").to_dict() if "LOTEFABRICACION" in alm_p0.columns else {}

    # Índices perfiles: variable 10
    artvar_p["ARTICULO"] = artvar_p["ARTICULO"].astype(str).str.strip()
    artvar10_p = set(artvar_p[pd.to_numeric(artvar_p["VARIABLE"], errors="coerce") == 10]["ARTICULO"])

    # Índices perfiles: características
    car_p["ARTICULO"]       = car_p["ARTICULO"].astype(str).str.strip()
    car_p["CARACTERISTICA"] = car_p["CARACTERISTICA"].astype(str).str.strip()
    car_p["VALOR"]          = car_p["VALOR"].astype(str).str.strip()
    carac_p = car_p.groupby("ARTICULO").apply(
        lambda g: g.set_index("CARACTERISTICA")["VALOR"].to_dict()
    ).to_dict()

    # Índice CLASE: código -> descripción en mayúsculas
    clases["CODIGO"] = clases["CODIGO"].astype(str).str.strip()
    clases["DENOMINACION"] = clases["DENOMINACION"].astype(str).str.upper().str.strip()
    desc_clase = clases.set_index("CODIGO")["DENOMINACION"].to_dict()

    inc = []
    for _, r in sub.iterrows():
        cod = r["CODIGO_ARTICULO"]
        estado = ni(r.get("ESTADO"))
        tipo = ni(r.get("TIPO"))
        # Artículos TIPO=2 + CLASE=3 (lacados/anodizados para compra): reglas propias
        es_lacado = (tipo == 2 and ni(r.get("CLASE")) == 3)
        # Clase 4 = anodizado especial: grupo asignado manualmente, no se valida
        es_clase4 = (ni(r.get("CLASE")) == 4)
        # Artículos CLASE=29 (embalaje): solo compra, sin tarifa 40, sin partida arancelaria, no en Excel JM
        es_embalaje = (ni(r.get("CLASE")) == 29)
        # Artículos CLASE=38/39 (catálogos y marketing): no en Excel JM, PVP debe ser igual al coste
        es_catalogo = (ni(r.get("CLASE")) in (38, 39))

        # Artículos que NO necesitan PVP en tarifa 40 (independiente del estado)
        _clase_pvp = ni(r.get("CLASE"))
        _necesita_pvp = not (
            es_lacado or es_embalaje or                      # tipo 2 cl.3 y cl.29
            (tipo == 1 and _clase_pvp in (0, 8)) or          # poliamidas, mecanizados
            (tipo == 2 and _clase_pvp == 4) or               # anodizados especiales
            tipo == 3 or                                     # ingresos extraordinarios
            (tipo == 4 and _clase_pvp == 8) or               # servicios curva/lacado
            (tipo == 50 and _clase_pvp == 41) or             # manipulación
            (tipo == 60 and _clase_pvp == 3)                 # lacado express
        )

        # 0) FAMILIA: debe coincidir con la indicada por JM en su Excel.
        fam_esp = ni(familia_jm.get(cod))
        fam_real = ni(r.get("FAMILIAPRODUCTO"))
        if cod in en_jm and fam_esp is not None and fam_real != fam_esp:
            extra = " (familia de PERFIL)" if fam_real in (1, 2) else ""
            inc.append((cod, "FAMILIA", "ERROR",
                        f"Familia debe ser {fam_esp} (según Excel JM). Encontrado: {fam_real}{extra}"))

        # 1) Campos fijos del maestro
        for campo, (esp, msg) in REGLAS_MAESTRO.items():
            val = r.get(campo)
            ok = (ni(val) == esp) if isinstance(esp, int) else (str(val).strip().upper() == str(esp).upper())
            if not ok:
                inc.append((cod, campo, "ERROR", f"{msg}. Encontrado: {val}"))

        # 2) CLASE: verificar que la descripción de la clase contiene alguna de las series indicadas por JM
        clase_cod = str(ni(r.get("CLASE"))) if ni(r.get("CLASE")) is not None else None
        serie_jm = clase_serie_jm.get(cod, "").strip().upper()
        if cod in en_jm and serie_jm and cod not in _TIPO_CLASE_UI_WHITELIST and cod not in _CLASE_WHITELIST:
            desc = desc_clase.get(clase_cod, "")
            # Las series en JM van separadas por "/" — basta con que al menos una aparezca en la descripción
            series = [s.strip() for s in serie_jm.split("/") if s.strip()]
            # Normalizar: quitar guiones y espacios extra para comparar
            def norm(t): return " ".join(t.replace("-", " ").split())
            # Alias de series (nombre comercial -> nombre en tabla de clases)
            ALIAS_SERIE = {"DOPLO": "PR77"}
            series_norm = [norm(ALIAS_SERIE.get(s, s)) for s in series]
            if desc and not any(s in norm(desc) for s in series_norm):
                inc.append((cod, "CLASE", "ERROR",
                            f"Clase {clase_cod} ('{desc}') no corresponde a la serie '{serie_jm}' indicada en Excel JM."))
            elif not desc and clase_cod is not None:
                inc.append((cod, "CLASE", "AVISO",
                            f"Clase {clase_cod} no encontrada en la tabla de clases."))

        # 3) Denominación base
        if not str(r.get("DENOMINACION", "")).strip() or str(r.get("DENOMINACION")).lower() == "nan":
            inc.append((cod, "DENOMINACION", "ERROR", "Falta denominación base."))

        # 3) MEDIDA
        ui, ue = ni(r.get("UNIDAD_INTERNA")), ni(r.get("UNIDAD_EXTERNA"))
        med_esp = 2 if (ui == 4 and ue == 4) else 1
        if ni(r.get("MEDIDA")) != med_esp:
            inc.append((cod, "MEDIDA", "ERROR", f"Medida debe ser {med_esp}. Encontrado: {ni(r.get('MEDIDA'))}"))

        # 3b) COEFICIENTE según caso de unidad
        coef = num(r.get("COEFICIENTE"))
        uart = num(r.get("UNIDADES_ARTICULO"))
        if ui == 9 and ue == 4:
            # Barras: coef debe ser el inverso de la unidad del artículo (1/longitud)
            esp_coef = round(1.0 / uart, 5) if uart else None
            if esp_coef is None:
                inc.append((cod, "COEFICIENTE", "ERROR", "Barra (9/4) sin unidad de artículo para calcular el coeficiente."))
            elif coef is None or abs(coef - esp_coef) > 0.0005:
                inc.append((cod, "COEFICIENTE", "ERROR",
                            f"Barra: coeficiente debe ser 1/{uart} = {esp_coef}. Encontrado: {coef}"))
        else:
            if coef != 1 and cod not in _COEF_MAESTRO_WHITELIST:
                inc.append((cod, "COEFICIENTE", "ERROR", f"Coeficiente debe ser 1. Encontrado: {coef}"))

        # 3c) UNIDADES_ARTICULO según caso de unidad
        if ui == 4 and ue == 4:
            # Metro lineal (rollos de junta/felpa): unidades artículo = LOTEHABITUAL del proveedor
            lote = lote_pr.get(cod)
            if lote is not None and lote > 0 and uart != lote:
                inc.append((cod, "UNIDADES_ARTICULO", "ERROR",
                            f"Metro lineal: unidades del artículo deben ser = lote habitual del proveedor ({lote}). Encontrado: {uart}"))
        elif ui == 9 and ue == 4:
            pass  # Barras: unidades del artículo = longitud (ya validado vía coeficiente)
        else:
            if uart != 1:
                inc.append((cod, "UNIDADES_ARTICULO", "ERROR", f"Unidades del artículo debe ser 1. Encontrado: {uart}"))

        # 4) Grupo según longitud y sufijo del código (lacados: grupo manual, no se valida):
        #    - 8 caracteres (sin acabado): grupo debe ser 0
        #    - >8 caracteres (con acabado): grupo según sufijo (01→12, 03→11, 04→10, 06→14, 05→16)
        grupo_val = ni(r.get("GRUPO"))
        d2_val = str(r.get("DENOMINACION_2", "") or "").strip().upper()
        if (es_lacado or es_clase4
                or (tipo == 91 and ni(r.get("CLASE")) == 3)
                or cod in _GRUPO_WHITELIST
                or (ni(r.get("CLASE")) == 27 and fam_real == 104)
                or (ni(r.get("CLASE")) == 33 and fam_real == 0)
                or (ni(r.get("CLASE")) == 36 and fam_real == 108)):
            pass  # grupo asignado manualmente en lacados, anodizado especial (clase 4),
                  # catálogos tipo 91 clase 3, whitelist, clase 27 fam 104, clase 33 fam 0,
                  # clase 35 fam 107, clase 36 fam 108
        else:
            # Sufijo estándar: posiciones 8-9 del código (01, 03, 04, 05, 06)
            sufijo = cod[8:] if len(cod) > 8 else ""
            g_sufijo = GRUPO_POR_SUFIJO.get(sufijo)
            # Excepción: prefijos con 11 dígitos (base 8 + letra dirección + 2 sufijo acabado)
            # Ej: PM402800D03 → sufijo "D03", últimos 2 = "03" = blanco
            if g_sufijo is None and len(sufijo) > 2 and cod.startswith(_PREFIJOS_SUFIJO_ULTIMOS2):
                g_sufijo = GRUPO_POR_SUFIJO.get(sufijo[-2:])
            # Excepción: A0063* con sufijos AM/BM/IM/NM → GRUPO 16 (PVD especial)
            _es_pvd_especial = False
            if g_sufijo is None and cod.startswith(_PREFIJOS_SUFIJO_PVD):
                g_sufijo = _SUFIJO_GRUPO_PVD.get(sufijo)
                if g_sufijo is not None:
                    _es_pvd_especial = True  # DEN2 es nombre de acabado específico, no "PVD"

            if g_sufijo is not None:
                # Código con sufijo de acabado reconocido → regla por sufijo
                if grupo_val == 0:
                    inc.append((cod, "GRUPO", "ERROR",
                                f"Código con sufijo de acabado: grupo no puede ser 0. Debe asignarse el grupo correspondiente."))
                elif grupo_val != g_sufijo:
                    inc.append((cod, "GRUPO", "ERROR",
                                f"Sufijo {sufijo} => grupo {g_sufijo}, pero GRUPO={grupo_val}."))
                # DENOMINACION_2 según grupo (omitir para A0063*: DEN2 es nombre de acabado específico)
                if not _es_pvd_especial:
                    esp2 = DENOM2_POR_GRUPO.get(g_sufijo)
                    if esp2 is not None and d2_val != esp2:
                        inc.append((cod, "DENOMINACION_2", "ERROR",
                                    f"Grupo {g_sufijo} => denominación 2 debe ser '{esp2}'. Encontrado: '{r.get('DENOMINACION_2')}'"))
            else:
                # Código sin sufijo estándar (longitud no estándar o sufijo no reconocido)
                # → el grupo se valida contra la DENOMINACION_2
                g_den2 = _GRUPO_POR_DEN2.get(d2_val)
                if g_den2 is not None:
                    # DEN2 es un acabado conocido: grupo debe coincidir
                    if grupo_val != g_den2:
                        inc.append((cod, "GRUPO", "ERROR",
                                    f"Denominación 2 '{d2_val}' => grupo esperado {g_den2}, "
                                    f"pero GRUPO={grupo_val}."))
                else:
                    # DEN2 no es un acabado estándar: grupo debe ser 0
                    if grupo_val != 0:
                        inc.append((cod, "GRUPO", "ERROR",
                                    f"Código sin sufijo de acabado y denominación 2 '{d2_val}' no reconocida: "
                                    f"grupo debe ser 0. Encontrado: {grupo_val}."))

        # 5) Imagen (característica 0) — los lacados (tipo 2, clase 3) y algunos artículos
        # de servicio nunca tienen imagen
        _clase = ni(r.get("CLASE"))
        _sin_imagen = (
            es_lacado
            or tipo == 3
            or (tipo == 4  and _clase == 8)
            or (tipo == 50 and _clase == 41 and fam_real == 0)
            or (tipo == 60 and _clase == 3  and fam_real == 0)
            or (tipo == 91 and _clase == 3  and fam_real == 0)
            or (tipo == 92 and fam_real == 107)
            or cod in _SIN_IMAGEN_WHITELIST
        )
        if not _sin_imagen and cod not in car0:
            inc.append((cod, "CARACTERISTICA_0", "ERROR", "Sin imagen en característica 0."))

        # 6) Proveedor principal + plazo + lote
        if cod not in con_principal:
            if _necesita_pvp or es_lacado:
                inc.append((cod, "PROVEEDOR", "ERROR", "Sin proveedor PRINCIPAL (PRINCIPAL=1) asignado."))
        else:
            prov_id = ni(proveedor_pr.get(cod))
            if es_lacado:
                # Lacados: plazo manual (no se valida), lote = 1, fórmula = 32
                lote = lote_pr.get(cod)
                if lote != 1:
                    inc.append((cod, "LOTE_HABITUAL", "ERROR",
                                f"Lacado (tipo 2, clase 3): lote habitual debe ser 1. Encontrado: {lote}"))
                formula = ni(formula_pr.get(cod))
                if formula != 32 and cod not in _ARTVAR_WHITELIST:
                    inc.append((cod, "FORMULA_PROVEEDOR", "ERROR",
                                f"Lacado (tipo 2, clase 3): fórmula debe ser 32. Encontrado: {formula}"))
            else:
                _clase_plazo = ni(r.get("CLASE"))
                plazo = plazo_pr.get(cod)
                if _clase_plazo in _PLAZO_CLASES_PENDIENTES:
                    pass  # clase 38 pendiente de ajuste, sin validar plazo
                elif not (plazo and plazo > 1):
                    inc.append((cod, "PLAZO_ENTREGA", "ERROR",
                                f"Plazo de entrega debe ser > 1 día. Encontrado: {plazo}."))
                else:
                    # Determinar plazo esperado: regla explícita > habitual estadístico
                    plazo_esp = None
                    if prov_id in _PLAZO_FIJO_PROV:
                        plazo_esp = _PLAZO_FIJO_PROV[prov_id]
                    elif prov_id in _PLAZO_PROV_POR_UI:
                        plazo_esp = _PLAZO_PROV_POR_UI[prov_id].get(ui)
                    elif prov_id in _PLAZO_PROV_POR_CLASE:
                        plazo_esp = _PLAZO_PROV_POR_CLASE[prov_id].get(_clase_plazo)
                    else:
                        hab = plazo_habitual.get(prov_id) if prov_id is not None else None
                        plazo_esp = hab if (hab is not None and hab > 1) else None
                    if plazo_esp is not None and plazo != plazo_esp:
                        inc.append((cod, "PLAZO_ENTREGA", "AVISO",
                                    f"Plazo {plazo} difiere del esperado para proveedor {prov_id} ({int(plazo_esp)} días). "
                                    "Revisar si es correcto."))
                if not (lote_pr.get(cod) and lote_pr[cod] > 0):
                    inc.append((cod, "LOTE_HABITUAL", "ERROR", "Lote habitual del proveedor principal a 0 o vacío."))
            # IVA proveedor: obligatorio si el proveedor es nacional (TIPO_CTA_COMPRA=0)
            tipo_cta = tipo_cta_prov.get(prov_id) if prov_id is not None else None
            if tipo_cta == 0:
                piva = piva_pr.get(cod)
                if piva is None or piva == 0:
                    inc.append((cod, "IVA_PROVEEDOR", "ERROR",
                                f"Proveedor nacional (España): IVA en artículo proveedor no informado o a 0. "
                                f"Encontrado: {piva}"))

        # 6b) Reglas adicionales artículo proveedor
        if cod in con_principal:
            # Ref. proveedor (CODIGO_PROVEEDOR) no puede estar vacía
            ref_pv = ref_prov_pr.get(cod, "")
            if not ref_pv or ref_pv.lower() in ("nan", "none", ""):
                inc.append((cod, "REF_PROVEEDOR", "ERROR",
                            "Referencia del proveedor (ref. proveedor) vacía o no informada."))
            # Coste bruto > 0
            cb = costebruto_pr.get(cod)
            if cb is None or cb <= 0:
                inc.append((cod, "COSTE_BRUTO", "ERROR",
                            f"Coste bruto del proveedor principal debe ser > 0. Encontrado: {cb}"))
            # Coeficiente proveedor = coeficiente maestro
            if cod not in _COEF_PROV_WHITELIST:
                coef_maestro = num(r.get("COEFICIENTE"))
                coef_prov = None
                if "COEFICIENTE" in pr.columns:
                    coef_prov = pd.to_numeric(pr.at[cod, "COEFICIENTE"], errors="coerce") if cod in pr.index else None
                if coef_prov is not None and coef_maestro is not None and coef_prov != coef_maestro:
                    inc.append((cod, "COEF_PROVEEDOR", "ERROR",
                                f"Coeficiente proveedor ({coef_prov}) debe ser igual al coeficiente maestro ({coef_maestro})."))
            # Lote habitual almacén = lote habitual proveedor
            lote_pv = lote_pr.get(cod)
            lote_al = lote_alm.get(cod)
            if lote_pv and lote_pv > 0 and lote_al is not None and lote_al != lote_pv and cod not in _LOTE_ALM_WHITELIST:
                inc.append((cod, "LOTE_ALMACEN", "ERROR",
                            f"Lote almacén ({lote_al}) debe coincidir con lote habitual proveedor ({lote_pv})."))

        # 6c) Artículo variable: UI ≠ UE => debe tener variable 10 (número de barras)
        if ui is not None and ue is not None and ui != ue and cod not in _ARTVAR_WHITELIST:
            if cod not in artvar_10:
                inc.append((cod, "ARTICULO_VARIABLE", "ERROR",
                            f"Unidad interna ({ui}) ≠ unidad externa ({ue}): "
                            "debe tener artículo variable 10 (número de barras)."))

        # 6d) Artículo asociado: tipo 92 grupo 14 => debe tener MINIMOACS, MINIMOACS2 o MINIMOACS3
        if tipo == 92 and grupo_val == 14 and not cod.startswith(_ARTASOC_SIN_MINIMO_PREFIJOS):
            asocs = asoc_por_art.get(cod, set())
            minimoacs = {"MINIMOACS", "MINIMOACS2", "MINIMOACS3"}
            if not (asocs & minimoacs):
                inc.append((cod, "ARTICULO_ASOCIADO", "ERROR",
                            "Tipo 92, grupo 14 (RAL estándar): debe tener artículo asociado "
                            "MINIMOACS, MINIMOACS2 o MINIMOACS3."))

        # 6e) Artículo asociado:
        #   - INTEGRACIONAUTO=1 (Comercial): obligatorio en todas las líneas de cualquier artículo
        #   - CANTIDAD=0: solo obligatorio en tipo 92 grupo 14 (mínimos RAL estándar)
        es_tipo92_g14 = (tipo == 92 and grupo_val == 14)
        for linea in ([] if cod in _ARTASOC_LINEAS_OK or cod.startswith(_ARTASOC_SIN_MINIMO_PREFIJOS) else asoc_lineas.get(cod, [])):
            art_asoc = linea.get("ARTICULOASOCIADO", "")
            cant = linea.get("CANTIDAD")
            integ = linea.get("INTEGRACIONAUTO")
            cant_val = float(cant) if cant is not None and str(cant) not in ("nan", "None") else None
            integ_val = int(integ) if integ is not None and str(integ) not in ("nan", "None") else None
            if es_tipo92_g14 and cant_val is not None and cant_val != 0:
                inc.append((cod, "ARTICULO_ASOCIADO", "ERROR",
                            f"Artículo asociado '{art_asoc}': tipo 92 grupo 14, cantidad debe ser 0. "
                            f"Encontrado: {cant_val}"))
            if integ_val != 1:
                inc.append((cod, "ARTICULO_ASOCIADO", "ERROR",
                            f"Artículo asociado '{art_asoc}': integración automática debe ser 'Comercial' (1). "
                            f"Encontrado: {integ_val}"))

        # 7) PVP / IVA tarifa 40 / Margen
        _clase_art = ni(r.get("CLASE"))
        if not _necesita_pvp:
            pass  # artículo sin tarifa de venta (lacados, embalajes, servicios…)
        elif es_catalogo:
            # Catálogos/marketing (clase 38/39): PVP obligatorio e igual al coste
            pvp = pvp40.get(cod)
            iva_t = iva_tar40.get(cod)
            if pvp is None:
                inc.append((cod, "PVP", "ERROR", "Sin PVP en tarifa 40."))
            if iva_t is None or iva_t == 0:
                inc.append((cod, "IVA_TARIFA", "ERROR",
                            f"IVA en tarifa 40 no informado o a 0. Encontrado: {iva_t}"))
            if cod not in _MARGEN_WHITELIST:
                coste = coste_pr.get(cod)
                if coste is not None and pvp is not None and pvp != coste:
                    inc.append((cod, "MARGEN", "ERROR",
                                f"Catálogo/marketing: PVP debe ser igual al coste. "
                                f"Coste: {round(coste, 4)}, PVP: {pvp}."))
        else:
            pvp = pvp40.get(cod)
            if pvp is None:
                inc.append((cod, "PVP", "ERROR", "Sin PVP en tarifa 40."))

            iva_t = iva_tar40.get(cod)
            if iva_t is None or iva_t == 0:
                inc.append((cod, "IVA_TARIFA", "ERROR",
                            f"IVA en tarifa 40 no informado o a 0. Encontrado: {iva_t}"))

            if cod not in _MARGEN_WHITELIST:
                coste, d = coste_pr.get(cod), divis.get(cod)
                # FW* clase 53: rango de margen especial 20-23%
                _es_fw53 = (cod.startswith("FW") and _clase_art == 53)
                if _es_fw53:
                    if coste is not None and pvp is not None and pvp > 0:
                        margen_fw = (pvp - coste) / pvp * 100
                        if not (_MARGEN_FW_CLASE53_MIN <= margen_fw <= _MARGEN_FW_CLASE53_MAX):
                            inc.append((cod, "MARGEN", "ERROR",
                                        f"FW clase 53: margen debe estar entre {_MARGEN_FW_CLASE53_MIN}% y "
                                        f"{_MARGEN_FW_CLASE53_MAX}%. Encontrado: {round(margen_fw, 1)}%. "
                                        f"Coste: {round(coste, 4)}, PVP: {pvp}."))
                elif cod not in en_jm:
                    # Sin divisor JM: regla simple >= 30%
                    if coste is not None and pvp is not None and pvp > 0:
                        margen_simple = (pvp - coste) / pvp * 100
                        if margen_simple < 30:
                            inc.append((cod, "MARGEN", "ERROR",
                                        f"Margen bruto {round(margen_simple, 1)}% inferior al mínimo del 30%. "
                                        f"Coste: {round(coste, 4)}, PVP tarifa 40: {pvp}."))
                elif coste is None:
                    pass
                elif d is None or d == 0:
                    inc.append((cod, "DIVISOR", "ERROR", "Divisor de accesorio vacío o 0 en el Excel de José María."))
                elif pvp is None:
                    pass
                else:
                    calc = round(coste / d, 2)
                    absdif = abs(calc - pvp)
                    pct = (absdif / pvp * 100) if pvp else 0
                    if absdif > TOL_MARGEN_EUR and pct > TOL_MARGEN_PCT:
                        if pvp < calc:
                            nivel, etiqueta = "ERROR", "vende por DEBAJO del PVP calculado (poco margen)"
                        else:
                            nivel, etiqueta = "AVISO", "PVP por encima del calculado (más margen)"
                        inc.append((cod, "MARGEN", nivel,
                                    f"{etiqueta}. Esperado {calc} (coste actual {round(coste,4)}/div {d}) vs real {pvp}. "
                                    f"Dif {round(calc-pvp,2)} ({round(pct,1)}%)."))

        # 8) ESTADO: lacados siempre 0; resto según Excel JM
        if es_lacado:
            if estado != 0:
                inc.append((cod, "ESTADO", "ERROR",
                            f"Lacado (tipo 2, clase 3): estado debe ser 0. Encontrado: {estado}"))
        else:
            estado_esp = ni(estado_jm.get(cod))
            if cod not in _ESTADO_WHITELIST and cod in en_jm and estado_esp is not None and estado != estado_esp:
                inc.append((cod, "ESTADO", "ERROR",
                            f"Estado debe ser {int(estado_esp)} (según Excel JM). Encontrado: {estado}"))

        # 8b) Stock mínimo (lacados no gestionan stock mínimo — son artículos de compra): cruzar Excel JM (col. MINIMO ACCESORIO) vs ALMACEN, según estado.
        #   - JM = 0  y almacén > 0                  -> ERROR (no debe tener mínimo)
        #   - JM > 0  y almacén > 0                  -> OK (el valor exacto se ajusta por ventas)
        #   - JM > 0  y almacén = 0 y estado(JM) 70  -> ERROR (en tarifa: debe tener mínimo)
        #   - JM > 0  y almacén = 0 y estado(JM) != 70 -> AVISO (aún no en tarifa)
        sm = stockmin.get(cod) if not es_lacado else None
        mj = minimo_jm.get(cod) if not es_lacado else None
        if mj is not None and not pd.isna(mj):  # celda vacía en Excel JM => no se valida
            sm_pos = sm is not None and sm > 0
            if mj == 0:
                if sm_pos:
                    inc.append((cod, "STOCK_MINIMO", "ERROR",
                                f"Excel JM indica mínimo 0 pero almacén tiene {sm}."))
            else:  # mj > 0
                if not sm_pos:
                    estado_esp_sm = ni(estado_jm.get(cod))
                    if estado_esp_sm == 70:
                        inc.append((cod, "STOCK_MINIMO", "ERROR",
                                    f"Estado 70 (en tarifa): debe tener stock mínimo en almacén. "
                                    f"Excel JM indica {mj} y almacén = {sm}."))
                    else:
                        inc.append((cod, "STOCK_MINIMO", "AVISO",
                                    f"Excel JM indica mínimo {mj} pero almacén no tiene mínimo (={sm})."))

        # 9) TIPO (accesorios). Reglas confirmadas con JM:
        #    - 0, 16, 91 NO deben existir en accesorios.
        #    - Con estructura que contiene CATALOGOINTERIORES => 92 (modular).
        #    - Con estructura que contiene OTRO catálogo (tipo 91, p.ej. MADERA/RALESTANDAR)
        #      => es de perfiles, no debería estar en accesorios.
        #    - Con estructura de códigos de compraventa => 15 (kit, solo accesorios).
        #    - Sin estructura => 1 (compraventa), 2 (lacado), 5 (muestra) o 60.
        comps = componentes.get(cod, set())
        cat_comps = comps & catalogos  # catálogos presentes en la estructura
        otros_cat = cat_comps - {"CATALOGOINTERIORES"}
        if cod in _TIPO_CLASE_UI_WHITELIST:
            pass
        elif tipo == 92 and not comps:
            inc.append((cod, "TIPO", "ERROR",
                        "Tipo 92 (modular): debe tener estructura con artículo base + CATALOGOINTERIORES. "
                        "No tiene ningún componente."))
        elif tipo in (0, 16, 91):
            # Excepción: tipo 91 familia 0 clase 3 o 4 (catálogos de acabado especial): correcto
            if not (tipo == 91 and fam_real == 0 and ni(r.get("CLASE")) in (3, 4)):
                inc.append((cod, "TIPO", "ERROR", f"Tipo {tipo} no debe existir en accesorios."))
        elif comps:
            if "CATALOGOINTERIORES" in comps:
                if tipo != 92:
                    inc.append((cod, "TIPO", "ERROR",
                                f"Tiene CATALOGOINTERIORES => debe ser tipo 92 (modular). Encontrado: {tipo}"))
                else:
                    # Tipo 92: tipo artículo modular debe ser 0 (CONTRA PEDIDO)
                    tam = ni(r.get("TIPOARTMODULAR"))
                    if tam != 0:
                        inc.append((cod, "TIPOARTMODULAR", "ERROR",
                                    f"Tipo 92 (modular): tipo artículo modular debe ser 'Contra pedido' (0). "
                                    f"Encontrado: {tam}"))
                    # Estructura correcta: artículo base + CATALOGOINTERIORES
                    # Base esperada: código completo si está en _TIPO92_BASE_FULLCODE, si no cod[:8]
                    _sin_base_ok = (cod in _TIPO92_SIN_BASE_CODIGOS
                                    or cod.startswith(_TIPO92_SIN_BASE_PREFIJOS))
                    if not _sin_base_ok:
                        base_esp = cod if cod in _TIPO92_BASE_FULLCODE else cod[:8]
                        comps_sin_cat = comps - {"CATALOGOINTERIORES"}
                        if base_esp not in comps:
                            inc.append((cod, "TIPO", "ERROR",
                                        f"Tipo 92 (modular): falta el artículo base '{base_esp}' en la estructura. "
                                        f"Componentes actuales: {sorted(comps_sin_cat) or '(ninguno)'}"))
                        extra = comps_sin_cat - {base_esp}
                        if extra:
                            inc.append((cod, "TIPO", "ERROR",
                                        f"Tipo 92 (modular): estructura tiene componentes inesperados además de "
                                        f"'{base_esp}' y CATALOGOINTERIORES: {sorted(extra)}"))
            elif otros_cat:
                inc.append((cod, "TIPO", "ERROR",
                            f"Estructura con catálogo de perfil {sorted(otros_cat)} (en accesorios solo vale CATALOGOINTERIORES). Tipo: {tipo}"))
            else:
                if tipo != 15:
                    inc.append((cod, "TIPO", "ERROR",
                                f"Estructura de códigos de compraventa => debe ser tipo 15 (kit). Encontrado: {tipo}"))
        else:
            # Tipo 3 (ingresos extraordinarios) familia 0 clase 0: correcto sin estructura
            # Tipo 4 clase 8: correcto sin estructura
            if ((tipo == 3 and fam_real == 0 and ni(r.get("CLASE")) == 0)
                    or (tipo == 4 and ni(r.get("CLASE")) == 8)
                    or (tipo == 50 and ni(r.get("CLASE")) == 41 and fam_real == 0)
                    or (tipo == 60 and ni(r.get("CLASE")) == 33 and fam_real == 0)
                    or cod in _TIPO_SIN_ESTRUCTURA_OK
                    or cod.startswith(_TIPO_SIN_ESTRUCTURA_PREFIJOS)):
                pass
            elif tipo not in (1, 2, 5, 60):
                inc.append((cod, "TIPO", "ERROR",
                            f"Sin estructura: tipo debe ser 1 (compraventa), 2 (lacado), 5 (muestra) o 60. Encontrado: {tipo}"))

        # 9b) Pestaña Más: PARTIDA ARANCELARIA y CONVERSIÓN M/KG
        partida = ni(r.get("PARTIDAARANCELARIA"))
        if es_lacado:
            # Lacados (tipo 2, clase 3): partida siempre debe ser 0
            if partida != 0:
                inc.append((cod, "PARTIDAARANCELARIA", "ERROR",
                            f"Lacado (tipo 2, clase 3): partida arancelaria debe ser 0. Encontrado: {partida}"))
        elif es_embalaje:
            pass  # Embalaje (clase 29): sin partida arancelaria, no se valida
        elif tipo == 3:
            # Ingresos extraordinarios (tipo 3): partida siempre debe ser 0
            if partida != 0:
                inc.append((cod, "PARTIDAARANCELARIA", "ERROR",
                            f"Ingreso extraordinario (tipo 3): partida arancelaria debe ser 0. Encontrado: {partida}"))
        else:
            partida_esp = ni(partida_jm.get(cod)) if cod in partida_jm else None
            if partida_esp is not None:
                if partida != partida_esp:
                    inc.append((cod, "PARTIDAARANCELARIA", "ERROR",
                                f"Partida arancelaria debe ser {partida_esp} (según Excel JM). Encontrado: {partida}"))
            else:
                if partida is None or partida == 0:
                    inc.append((cod, "PARTIDAARANCELARIA", "ERROR",
                                f"Partida arancelaria no puede ser 0 o vacía. Encontrado: {partida}"))
        conv_mkg = num(r.get("CONVERSION_M_KG"))
        if fam_real in (1, 2) and (conv_mkg is None or conv_mkg <= 0):
            inc.append((cod, "CONVERSION_M_KG", "ERROR",
                        f"Familia {fam_real} (perfil): coeficiente m/kg debe ser > 0. Encontrado: {conv_mkg}"))

        # 9c) Pestaña Mas2: checks
        vars_check = ni(r.get("GESTIONARVARIABLES"))
        subcont = ni(r.get("ESSUBCONTRATACIO"))
        eskit = ni(r.get("ESKIT"))
        calc_formula = ni(r.get("CALCULOCANTFORMULA"))

        # GESTIONARVARIABLES: debe estar marcado cuando unidad externa = 9 (barras)
        if ue == 9 and vars_check != 1:
            inc.append((cod, "GESTIONARVARIABLES", "ERROR",
                        "Unidad externa 9 (barras): 'Gestionar variables' debe estar marcado "
                        f"(permite indicar nº de barras en artículos variables). Encontrado: {vars_check}"))

        # ESSUBCONTRATACIO: debe estar marcado cuando TIPO = 92 (modular)
        if tipo == 92 and subcont != 1:
            inc.append((cod, "ESSUBCONTRATACIO", "ERROR",
                        f"Tipo 92 (modular): 'Es subcontratación' debe estar marcado. Encontrado: {subcont}"))

        # ESKIT: siempre debe estar vacío (nunca se usa en accesorios)
        if eskit == 1:
            inc.append((cod, "ESKIT_CHECK", "ERROR",
                        "El check 'Es un artículo kit' no debe estar marcado en accesorios."))

        # CALCULOCANTFORMULA debe estar marcado (=1) si unidad interna = 9
        if ui == 9 and calc_formula != 1 and cod not in _ARTVAR_WHITELIST:
            inc.append((cod, "CALCULOCANTFORMULA", "ERROR",
                        "Unidad interna 9 (barras): 'Cálculo cantidad en documentos comerciales por fórmula' "
                        f"debe estar marcado. Encontrado: {calc_formula}"))

        # NOGESTIONASTOCK: solo los artículos de la lista blanca pueden tenerlo marcado
        nostock = ni(r.get("NOGESTIONASTOCK"))
        en_whitelist = cod in _NOSTOCK_WHITELIST
        if nostock == 1 and not en_whitelist:
            inc.append((cod, "NOGESTIONASTOCK", "ERROR",
                        "No gestiona stock marcado pero este artículo no está en la lista de artículos "
                        "exentos de inventario. Revisar si es correcto o desmarcar."))
        elif nostock != 1 and en_whitelist:
            inc.append((cod, "NOGESTIONASTOCK", "ERROR",
                        "Este artículo debe tener 'No gestiona inventario' marcado (está en la lista de exentos)."))

        # 10) Coherencia de unidades MAESTRO <-> PROVEEDOR <-> TARIFABLE
        # 10a) TIPO 2 (lacado/anodizado) <=> UNIDAD_INTERNA 5 (m2)
        if cod not in _TIPO_CLASE_UI_WHITELIST:
            if tipo == 2 and ui != 5:
                inc.append((cod, "UNIDAD/TIPO", "ERROR", f"Tipo 2 (lacado) debe tener unidad interna 5. Encontrado: {ui}"))
            if ui == 5 and tipo != 2:
                inc.append((cod, "UNIDAD/TIPO", "ERROR", f"Unidad interna 5 (m2) debe ser tipo 2. Encontrado tipo: {tipo}"))
        # 10b) externa del maestro = unidad del proveedor principal
        uep = uext_pr.get(cod)
        if uep is not None and ue is not None and uep != ue and cod not in _UNIDAD_PROV_WHITELIST:
            inc.append((cod, "UNIDAD_PROVEEDOR", "ERROR",
                        f"Unidad externa maestro={ue} pero proveedor principal={int(uep)}."))
        # 10c) unidad de venta tarifa 40 coherente con interna (1->1, 9->4, 4->4, 5->5)
        venta_esp = {1: 1, 9: 4, 4: 4, 5: 5}.get(ui)
        uvt = uventa_tar.get(cod)
        if venta_esp is not None and uvt is not None and uvt != venta_esp:
            inc.append((cod, "UNIDAD_VENTA", "ERROR",
                        f"Interna {ui} => unidad de venta tarifa esperada {venta_esp}. Encontrado: {int(uvt)}."))

    # ── FAMILIA 105: chapas y paneles ──────────────────────────────────────────
    def _norm_acc(s):
        return "".join(c for c in unicodedata.normalize("NFD", str(s).upper().strip())
                       if unicodedata.category(c) != "Mn")

    for _, r in sub_105.iterrows():
        cod = r["CODIGO_ARTICULO"]
        tipo = ni(r.get("TIPO"))
        clase = ni(r.get("CLASE"))
        grupo_val = ni(r.get("GRUPO"))
        artbase = str(r.get("ARTICULOBASE", "") or "").strip()
        den2 = str(r.get("DENOMINACION_2", "") or "").strip()
        nostock = ni(r.get("NOGESTIONASTOCK"))

        # Artículos de servicio dentro de familia 105 (incrementos, complementos…):
        # clase 0 → validación reducida: proveedor, coste, PVP y partida arancelaria = 0.
        if clase == 0:
            if cod not in _F105_BAJO_PRESUPUESTO:
                if cod not in con_principal:
                    inc.append((cod, "PROVEEDOR", "ERROR",
                                "Complemento F105 (clase 0): sin proveedor PRINCIPAL asignado."))
                else:
                    if not (coste_pr.get(cod) and coste_pr[cod] > 0):
                        inc.append((cod, "PROVEEDOR", "ERROR",
                                    "Complemento F105 (clase 0): coste del proveedor principal a 0 o vacío."))
                pvp_c0 = pvp40.get(cod)
                if pvp_c0 is None:
                    inc.append((cod, "PVP", "ERROR",
                                "Complemento F105 (clase 0): sin PVP en tarifa 40."))
            if cod not in _F105_BAJO_PRESUPUESTO:
                partida_c0 = ni(r.get("PARTIDAARANCELARIA"))
                if partida_c0 != 0:
                    inc.append((cod, "PARTIDAARANCELARIA", "ERROR",
                                f"Complemento F105 (clase 0): partida arancelaria debe ser 0. "
                                f"Encontrado: {partida_c0}"))
            continue

        # CLASE: 31 (chapa), 54 (panel), 56 (puerta panelada)
        if clase not in _F105_CLASES_VALIDAS:
            inc.append((cod, "CLASE", "ERROR",
                        f"Familia 105: clase debe ser 31 (chapa), 54 (panel) o 56 (puerta panelada). "
                        f"Encontrado: {clase}"))

        if tipo == 60:
            # Variante de acabado: DENOMINACION_2 no vacía, ARTICULOBASE relleno,
            # GRUPO igual al del artículo base, sin inventario bloqueado.
            # XDECOR*/XTA*: DEN2 = dimensiones+acabado (libre), solo se valida que no esté vacía.
            if not den2 or den2.lower() in ("nan", "none"):
                inc.append((cod, "DENOMINACION_2", "ERROR",
                            "Tipo 60 (variante acabado): denominación 2 no puede estar vacía."))
            if not artbase or artbase.lower() in ("nan", "none"):
                if clase != 31:
                    pass  # Clase 54 (panel) y 56 (puerta): no tienen artículo padre, es correcto
                else:
                    # Clase 31 (chapa): debe tener ARTICULOBASE. Intentar derivarlo de la estructura.
                    comps_t60 = componentes.get(cod, set())
                    _padres_f105 = _f105_tipo1 | set(_f105_grupo_de_92.keys())
                    padre_sug = next((c for c in comps_t60 if c in _padres_f105), None)
                    if padre_sug:
                        inc.append((cod, "ARTICULOBASE", "AVISO",
                                    f"Artículo base vacío. Sugerido: '{padre_sug}' "
                                    "(encontrado en estructura). Rellenar en Geinfor."))
                    else:
                        inc.append((cod, "ARTICULOBASE", "ERROR",
                                    "Artículo base vacío y no se puede derivar de la estructura. "
                                    "Revisar en Geinfor."))
            else:
                # Solo se valida coherencia de GRUPO cuando ARTICULOBASE es tipo 92
                # (tiene catálogo de acabado). Si es tipo 1 (base en bruto), el GRUPO
                # del tipo 60 viene del componente de acabado en la estructura, no del base.
                grupo_base = _f105_grupo_de_92.get(artbase)
                if grupo_base is not None and grupo_val != grupo_base:
                    inc.append((cod, "GRUPO", "ERROR",
                                f"Tipo 60: grupo {grupo_val} no coincide con el del artículo base "
                                f"'{artbase}' (grupo {grupo_base})."))
            if nostock == 1:
                inc.append((cod, "NOGESTIONASTOCK", "ERROR",
                            "Tipo 60 (chapa/panel): debe gestionar stock. "
                            "'No gestiona inventario' debe estar desmarcado."))

        elif tipo == 92:
            # Artículo padre: ARTICULOBASE vacío, estructura con catálogo tipo 91.
            # Clase 31 (chapa): DENOMINACION_2 = nombre del tipo de catálogo, GRUPO = grupo del catálogo.
            # Clase 54/56 (panel/puerta): usan CATALOGOINTERIORES (grupo 0) igual que accesorios
            #   modulares; DENOMINACION_2 = dimensiones, GRUPO independiente del catálogo.
            if artbase and artbase.lower() not in ("nan", "none"):
                inc.append((cod, "ARTICULOBASE", "ERROR",
                            f"Tipo 92 (artículo padre): artículo base debe estar vacío. "
                            f"Encontrado: '{artbase}'"))
            if clase == 31 and not cod.startswith(_F105_CON_IMAGEN_PREFIJOS):
                # Chapas estándar: DENOMINACION_2 debe coincidir con el nombre del catálogo según grupo.
                # Excepción: XDECOR* y XTA* llevan dimensiones+acabado en DEN2, no nombre de catálogo.
                if grupo_val in _F105_DEN2_POR_GRUPO:
                    esp_den2 = _F105_DEN2_POR_GRUPO[grupo_val]
                    if _norm_acc(den2) != _norm_acc(esp_den2):
                        inc.append((cod, "DENOMINACION_2", "ERROR",
                                    f"Tipo 92 grupo {grupo_val}: denominación 2 debe ser '{esp_den2}'. "
                                    f"Encontrado: '{den2}'"))
                elif not den2 or den2.lower() in ("nan", "none"):
                    inc.append((cod, "DENOMINACION_2", "ERROR",
                                "Tipo 92 (artículo padre): denominación 2 no puede estar vacía."))
            else:
                # Paneles/puertas: DENOMINACION_2 = dimensiones, no se valida su contenido
                if not den2 or den2.lower() in ("nan", "none"):
                    inc.append((cod, "DENOMINACION_2", "ERROR",
                                "Tipo 92 (panel/puerta padre): denominación 2 no puede estar vacía."))
            comps_92 = componentes.get(cod, set())
            cats_92 = comps_92 & catalogos
            if not cats_92:
                inc.append((cod, "TIPO", "ERROR",
                            "Tipo 92 (artículo padre): estructura vacía o sin catálogo de acabado "
                            "(artículo tipo 91)."))
            elif clase == 31:
                # Chapas: el grupo del catálogo debe coincidir con el grupo del artículo
                for cat in cats_92:
                    grupo_cat = _f105_grupo_de_cat.get(cat)
                    if grupo_cat is not None and grupo_cat != grupo_val:
                        inc.append((cod, "GRUPO", "ERROR",
                                    f"Tipo 92: catálogo '{cat}' (grupo {grupo_cat}) no coincide con "
                                    f"el grupo del artículo ({grupo_val})."))
            if nostock == 1:
                inc.append((cod, "NOGESTIONASTOCK", "ERROR",
                            "Tipo 92 (chapa/panel padre): debe gestionar stock."))

        elif tipo == 1:
            # Chapa/panel base en bruto (sin acabado): GRUPO debe ser 0.
            # Excepciones: artículos que ya vienen con acabado de fábrica.
            if grupo_val != 0 and cod not in _F105_TIPO1_CON_ACABADO:
                inc.append((cod, "GRUPO", "ERROR",
                            f"Tipo 1 (chapa/panel base): artículo en bruto sin acabado, "
                            f"grupo debe ser 0. Encontrado: {grupo_val}"))
            if nostock == 1:
                inc.append((cod, "NOGESTIONASTOCK", "ERROR",
                            "Tipo 1 (chapa/panel base): debe gestionar stock."))

        else:
            inc.append((cod, "TIPO", "ERROR",
                        f"Familia 105: tipo debe ser 60 (variante acabado), 92 (artículo padre) "
                        f"o 1 (chapa base). Encontrado: {tipo}"))

        # ── Comprobaciones comunes a TODOS los subtipos de F105 (clase != 0) ──

        # Imagen (característica 0):
        # - Clase 31 (chapas): NO debe tener imagen, salvo excepciones conocidas
        # - Clase 54/56 (paneles/puertas): SÍ debe tener imagen
        if clase == 31:
            tiene_excepcion = cod in _F105_CON_IMAGEN or cod.startswith(_F105_CON_IMAGEN_PREFIJOS)
            if cod in car0 and not tiene_excepcion:
                inc.append((cod, "CARACTERISTICA_0", "ERROR",
                            "Chapa (clase 31): no debe tener imagen en característica 0."))
        elif clase in (54, 56):
            if cod not in car0:
                inc.append((cod, "CARACTERISTICA_0", "ERROR",
                            "Panel/puerta (clase 54/56): debe tener imagen en característica 0."))

        # Partida arancelaria: clase 56 (puertas paneladas) = 18, resto (31/54) = 10
        partida_f105 = ni(r.get("PARTIDAARANCELARIA"))
        partida_esp = 18 if clase == 56 else 10
        if partida_f105 != partida_esp:
            inc.append((cod, "PARTIDAARANCELARIA", "ERROR",
                        f"Familia 105: partida arancelaria debe ser {partida_esp} (clase {clase}). "
                        f"Encontrado: {partida_f105}"))

        # Artículo 92 actúa como "artículo padre modular": no tiene PVP propio ni proveedor directo.
        # Tipos 60 y 1 sí deben tener tarifa 40 + IVA + PVP.
        if tipo in (60, 1):
            # Proveedor principal: obligatorio cuando el artículo no tiene estructura propia
            # (es subcontratado). Si tiene estructura, el coste viene de los componentes.
            comps_f105 = componentes.get(cod, set())
            tiene_estructura = bool(comps_f105)
            if not tiene_estructura:
                if cod not in con_principal:
                    inc.append((cod, "PROVEEDOR", "ERROR",
                                "Sin proveedor PRINCIPAL asignado (artículo sin estructura: "
                                "necesita proveedor de subcontratación)."))
                else:
                    plazo = plazo_pr.get(cod)
                    if not (plazo and plazo > 1):
                        inc.append((cod, "PLAZO_ENTREGA", "ERROR",
                                    f"Plazo de entrega debe ser > 1 día. Encontrado: {plazo}."))
                    if not (lote_pr.get(cod) and lote_pr[cod] > 0):
                        inc.append((cod, "LOTE_HABITUAL", "ERROR",
                                    "Lote habitual del proveedor principal a 0 o vacío."))
                    prov_id_f = ni(proveedor_pr.get(cod))
                    tipo_cta_f = tipo_cta_prov.get(prov_id_f) if prov_id_f is not None else None
                    if tipo_cta_f == 0:
                        piva_f = piva_pr.get(cod)
                        if piva_f is None or piva_f == 0:
                            inc.append((cod, "IVA_PROVEEDOR", "ERROR",
                                        "Proveedor nacional: IVA en artículo proveedor no informado o a 0."))

            # Tarifa 40: PVP e IVA obligatorios
            pvp_f = pvp40.get(cod)
            if pvp_f is None:
                inc.append((cod, "PVP", "ERROR", "Sin PVP en tarifa 40."))
            iva_f = iva_tar40.get(cod)
            if iva_f is None or iva_f == 0:
                inc.append((cod, "IVA_TARIFA", "ERROR",
                            f"IVA en tarifa 40 no informado o a 0. Encontrado: {iva_f}"))

            # Margen: familia 105 no usa divisor JM — se valida que el margen bruto >= 30%
            coste_f = coste_pr.get(cod) if not tiene_estructura else num(r.get("COSTE"))
            if coste_f is not None and pvp_f is not None and pvp_f > 0:
                margen_f = (pvp_f - coste_f) / pvp_f * 100
                if margen_f < 30:
                    inc.append((cod, "MARGEN", "ERROR",
                                f"Margen bruto {round(margen_f, 1)}% inferior al mínimo del 30%. "
                                f"Coste: {round(coste_f, 4)}, PVP tarifa 40: {pvp_f}."))

            # Stock mínimo: mismo criterio que accesorios
            sm_f = stockmin.get(cod)
            mj_f = minimo_jm.get(cod)
            if mj_f is not None and not pd.isna(mj_f):
                sm_pos_f = sm_f is not None and sm_f > 0
                if mj_f == 0:
                    if sm_pos_f:
                        inc.append((cod, "STOCK_MINIMO", "ERROR",
                                    f"Excel JM indica mínimo 0 pero almacén tiene {sm_f}."))
                else:
                    if not sm_pos_f:
                        est_jm_f = ni(estado_jm.get(cod))
                        if est_jm_f == 70:
                            inc.append((cod, "STOCK_MINIMO", "ERROR",
                                        f"Estado 70 (en tarifa): debe tener stock mínimo. "
                                        f"Excel JM indica {mj_f} y almacén = {sm_f}."))
                        else:
                            inc.append((cod, "STOCK_MINIMO", "AVISO",
                                        f"Excel JM indica mínimo {mj_f} pero almacén no tiene mínimo."))

            # ESTADO: según Excel JM
            estado_f = ni(r.get("ESTADO"))
            estado_esp_f = ni(estado_jm.get(cod))
            if cod in en_jm and estado_esp_f is not None and estado_f != estado_esp_f:
                inc.append((cod, "ESTADO", "ERROR",
                            f"Estado debe ser {int(estado_esp_f)} (según Excel JM). "
                            f"Encontrado: {estado_f}"))

    # ── FIN FAMILIA 105 ────────────────────────────────────────────────────────

    # ── PERFILES F1 / F2 ──────────────────────────────────────────────────────
    TARIFA_PERF = 40
    UD_VENTA_PERF = 4       # MT
    IVA_TARIFA_PERF = 21
    UD_INTERNA_PERF = 9
    MEDIDA_PERF = 1
    PARTIDA_PERF = 1
    ENSAMBLADO_PREFIJO = "ENSAM"   # código artículo ENSAMBLADO en estructura

    for _, rp in sub_p.iterrows():
        cod = str(rp["CODIGO_ARTICULO"]).strip()
        fam_p = ni(rp.get("FAMILIAPRODUCTO"))
        grp_p = ni(rp.get("GRUPO"))
        if grp_p is None:
            inc.append((cod, "GRUPO", "ERROR", "Sin grupo definido.")); continue
        grp_p = int(grp_p)
        base_p = str(rp.get("ARTICULOBASE", "") or "").strip()
        tipo_p = ni(rp.get("TIPO"))
        tipmod_p = ni(rp.get("TIPOARTMODULAR"))
        ud_int_p = ni(rp.get("UNIDAD_INTERNA"))
        ud_ext_p = ni(rp.get("UNIDAD_EXTERNA"))
        medida_p = ni(rp.get("MEDIDA"))
        coef_p   = num(rp.get("COEFICIENTE"))
        ud_art_p = num(rp.get("UNIDADES_ARTICULO"))
        estado_p = ni(rp.get("ESTADO"))
        gest_var  = ni(rp.get("GESTIONARVARIABLES"))
        es_kit    = ni(rp.get("ESKIT_CHECK"))
        nogest    = ni(rp.get("NOGESTIONASTOCK"))
        essubcon  = ni(rp.get("ESSUBCONTRATACIO"))
        calcform  = ni(rp.get("CALCULOCANTFORMULA"))
        den_p    = str(rp.get("DENOMINACION", "") or "").strip()

        # Datos Excel costes
        # La clave del Excel es la referencia BASE (8 chars) que coincide con el artículo base
        clave_cp = base_p[:8] if base_p else cod[:8]
        cp = costes_p.get(clave_cp) or costes_p.get(cod[:8]) or {}
        peso_m    = num(cp.get("peso"))
        perimetro = num(cp.get("perimetro"))
        perimetro_ext = num(cp.get("perimetro_ext"))
        caras_cp  = num(cp.get("caras"))
        pvps_cp   = cp.get("pvps", {})   # dict {grupo: pvp}
        costes_cp = cp.get("costes", {}) # dict {grupo: coste}

        # ── TIPO / TIPOARTMODULAR ──────────────────────────────────────────────
        # len(cod) == len(base)+2 → artículo con grupo (sin acabado) → tipo 92
        # len(cod) == len(base)+5 → artículo con acabado específico  → tipo 60
        len_base = len(base_p) if base_p else 0
        # ARTICULOBASE guarda el código base+grupo (p.ej. CF22411506).
        # Si cod == base → artículo de grupo sin acabado → tipo 92
        # Si cod > base  → tiene sufijo de acabado       → tipo 60
        con_acabado = len_base > 0 and len(cod) > len_base

        if grp_p == 0:
            tipo_esp = 1 if fam_p == 1 else 16
        elif grp_p in (1, 3):
            tipo_esp = 60
        else:  # grupos 2/4/5/6/7/9
            tipo_esp = 60 if con_acabado else 92

        if tipo_p != tipo_esp:
            inc.append((cod, "TIPO", "ERROR",
                        f"Debe ser {tipo_esp} (F{fam_p} grupo {grp_p}"
                        f"{', con acabado' if con_acabado else ''}). Encontrado: {tipo_p}"))

        tipmod_esp = 1 if grp_p in (2, 4, 5, 6, 7, 9) else 0
        if tipmod_p != tipmod_esp:
            inc.append((cod, "TIPOARTMODULAR", "ERROR",
                        f"Debe ser {tipmod_esp}. Encontrado: {tipmod_p}"))

        # ── UNIDADES ──────────────────────────────────────────────────────────
        if ud_int_p != UD_INTERNA_PERF:
            inc.append((cod, "UNIDAD/TIPO", "ERROR",
                        f"Unidad interna debe ser 9. Encontrado: {ud_int_p}"))
        ud_ext_esp = 2 if (fam_p == 1 and grp_p == 0) else 4
        if ud_ext_p != ud_ext_esp:
            inc.append((cod, "UNIDAD/TIPO", "ERROR",
                        f"Unidad externa debe ser {ud_ext_esp} ({'KG' if fam_p==1 else 'MT'}). Encontrado: {ud_ext_p}"))
        if medida_p != MEDIDA_PERF:
            inc.append((cod, "MEDIDA", "ERROR",
                        f"Medida debe ser 1. Encontrado: {medida_p}"))

        # ── COEFICIENTE ────────────────────────────────────────────────────────
        if ud_art_p and ud_art_p > 0:
            if fam_p == 1 and grp_p == 0 and peso_m and peso_m > 0:
                coef_esp = 1.0 / (peso_m * ud_art_p)  # F1 grupo 0: 1/(peso×ud_art)
            else:
                coef_esp = 1.0 / ud_art_p  # F1 grupos 1-9 y F2: 1/ud_art
            if coef_p is None or abs(coef_p - coef_esp) > 0.0001:
                inc.append((cod, "COEFICIENTE", "ERROR",
                            f"Debe ser {coef_esp:.6f} (1/{'peso×ud' if fam_p==1 else 'ud_art'}). "
                            f"Encontrado: {coef_p}"))

        # ── PESTAÑA MÁS ───────────────────────────────────────────────────────
        partida_val = ni(rp.get("PARTIDAARANCELARIA"))
        if partida_val != PARTIDA_PERF:
            inc.append((cod, "PARTIDAARANCELARIA", "ERROR",
                        f"Debe ser 1. Encontrado: {partida_val}"))
        conv_mkg = num(rp.get("CONVERSION_M_KG"))
        if peso_m is not None and conv_mkg is not None:
            if abs(conv_mkg - peso_m) > 0.001:
                inc.append((cod, "CONVERSION_M_KG", "ERROR",
                            f"Debe ser {peso_m} (peso kg/m del Excel costes). Encontrado: {conv_mkg}"))
        elif peso_m is not None and conv_mkg is None:
            inc.append((cod, "CONVERSION_M_KG", "ERROR",
                        f"Sin valor. Debe ser {peso_m}"))

        # F2: PESO_ARTICULO y PESO_BULTO
        if fam_p == 2:
            peso_art = num(rp.get("PESO_ARTICULO"))
            if peso_m is not None and (peso_art is None or abs(peso_art - peso_m) > 0.001):
                inc.append((cod, "PESO_ARTICULO", "ERROR",
                            f"Debe ser {peso_m}. Encontrado: {peso_art}"))
            if peso_m and ud_art_p:
                peso_bulto_esp = peso_m * ud_art_p
                peso_bulto = num(rp.get("PESO_BULTO"))
                if peso_bulto is None or abs(peso_bulto - peso_bulto_esp) > 0.001:
                    inc.append((cod, "PESO_BULTO", "ERROR",
                                f"Debe ser {peso_bulto_esp:.4f} (peso×ud_art). Encontrado: {peso_bulto}"))

        # ── CARACTERÍSTICAS ────────────────────────────────────────────────────
        caracs = carac_p.get(cod, {})

        def num_carac(v):
            """Parsea valor de característica admitiendo coma como separador decimal."""
            if v is None:
                return None
            s = str(v).strip().replace(",", ".")
            try:
                return float(s)
            except Exception:
                return None

        # Nº CARAS (100)
        caras_art = num_carac(caracs.get("100"))
        if caras_cp is not None and (caras_art is None or abs(caras_art - caras_cp) > 0.01):
            inc.append((cod, "CARACTERISTICA_100", "ERROR",
                        f"Nº CARAS debe ser {caras_cp}. Encontrado: {caracs.get('100')}"))
        # PERIMETRO TOTAL (200)
        perim_art = num_carac(caracs.get("200"))
        if perimetro is not None and (perim_art is None or abs(perim_art - perimetro) > 0.01):
            inc.append((cod, "CARACTERISTICA_200", "ERROR",
                        f"PERIMETRO TOTAL debe ser {perimetro}. Encontrado: {caracs.get('200')}"))
        # PERIMETRO EXTERIOR (201)
        perim_ext_art = num_carac(caracs.get("201"))
        if perimetro_ext is not None and (perim_ext_art is None or abs(perim_ext_art - perimetro_ext) > 0.01):
            inc.append((cod, "CARACTERISTICA_201", "ERROR",
                        f"PERIMETRO EXTERIOR debe ser {perimetro_ext}. Encontrado: {caracs.get('201')}"))

        # ── ARTÍCULO VARIABLE 10 ───────────────────────────────────────────────
        if cod not in artvar10_p:
            inc.append((cod, "ARTICULO_VARIABLE", "ERROR",
                        "Falta VARIABLE 10 (NUMERO DE BARRAS)."))

        # ── PROVEEDOR (solo grupo 0) ───────────────────────────────────────────
        if grp_p == 0:
            if cod not in con_principal_p:
                inc.append((cod, "PROVEEDOR", "ERROR",
                            "Sin proveedor PRINCIPAL definido."))
            else:
                # Unidad proveedor = unidad externa
                uext_prov = ni(uext_pr_p.get(cod))
                if uext_prov != ud_ext_esp:
                    inc.append((cod, "UNIDAD_PROVEEDOR", "ERROR",
                                f"Unidad proveedor debe ser {ud_ext_esp}. Encontrado: {uext_prov}"))
                # Coeficiente proveedor = coeficiente general
                coef_prov = num(coef_pr_p.get(cod))
                if coef_p is not None and coef_prov is not None and abs(coef_prov - coef_p) > 0.0001:
                    inc.append((cod, "COEF_PROVEEDOR", "ERROR",
                                f"Coeficiente proveedor ({coef_prov:.6f}) debe coincidir con "
                                f"coeficiente general ({coef_p:.6f})."))
                # Lote habitual: viene de JM
                lote_jm_val = _jm_lote.get(cod)
                lote_prov   = num(lote_pr_p.get(cod))
                if lote_jm_val is not None and lote_prov is not None:
                    if abs(lote_prov - lote_jm_val) > 0.001:
                        inc.append((cod, "LOTE_HABITUAL", "ERROR",
                                    f"Lote proveedor {lote_prov} ≠ Excel JM {lote_jm_val}."))
                # IVA: TIPO_CTA_COMPRA=0 → nacional (IVA=21), resto → internacional (IVA=0)
                prov_id_p = num(proveedor_pr_p.get(cod))
                tipo_cta_p = tipo_cta_prov.get(prov_id_p) if prov_id_p is not None else None
                if tipo_cta_p == 0:
                    iva_prov = ni(piva_pr_p.get(cod))
                    if iva_prov is None or iva_prov == 0:
                        inc.append((cod, "IVA_PROVEEDOR", "ERROR",
                                    f"Proveedor nacional: IVA debe ser 21. Encontrado: {iva_prov}"))

        # ── TARIFA 40 ──────────────────────────────────────────────────────────
        pvp_art = num(pvp40_p.get(cod))
        uv_art  = ni(uventa_p.get(cod))
        iva_t   = ni(iva_tar40_p.get(cod))
        coef_t  = num(coef_tar_p.get(cod))

        if uv_art != UD_VENTA_PERF:
            inc.append((cod, "UNIDAD_VENTA", "ERROR",
                        f"Unidad venta tarifa 40 debe ser 4 (MT). Encontrado: {uv_art}"))
        if iva_t != IVA_TARIFA_PERF:
            inc.append((cod, "IVA_TARIFA", "ERROR",
                        f"IVA tarifa 40 debe ser 21. Encontrado: {iva_t}"))
        if ud_art_p and ud_art_p > 0:
            coef_t_esp = 1.0 / ud_art_p
            if coef_t is None or abs(coef_t - coef_t_esp) > 0.0001:
                inc.append((cod, "COEFICIENTE", "ERROR",
                            f"Coeficiente tarifa 40 debe ser {coef_t_esp:.6f} (1/ud_art). "
                            f"Encontrado: {coef_t}"))

        # PVP tarifa 40
        _finish_key = _GRUPO_FINISH.get(grp_p)
        pvp_cp = num(pvps_cp.get(_finish_key)) if _finish_key else None
        if grp_p == 8:
            # Bicolor: debe ser ≥ PVP grupo 9 × 1.05
            pvp_g9 = num(pvp40_p.get(base_p + "09") if base_p else None)
            if pvp_g9 is None:
                # buscar artículo G9 con mismo base
                for k in pvp40_p:
                    if k[:8] == cod[:8] and k.endswith("09"):
                        pvp_g9 = num(pvp40_p[k]); break
            if pvp_g9 is not None and pvp_art is not None:
                pvp_esp_8 = round(pvp_g9 * 1.05, 6)
                if abs(pvp_art - pvp_esp_8) > 0.01:
                    inc.append((cod, "PVP", "ERROR",
                                f"PVP bicolor debe ser PVP grupo 9 × 1.05 = {pvp_esp_8:.4f}. "
                                f"Encontrado: {pvp_art}"))
        elif pvp_cp is not None and pvp_art is not None:
            if abs(pvp_art - pvp_cp) > 0.01:
                inc.append((cod, "PVP", "ERROR",
                            f"PVP debe ser {pvp_cp} (Excel costes grupo {grp_p}). "
                            f"Encontrado: {pvp_art}"))
        elif pvp_art is None:
            inc.append((cod, "PVP", "ERROR",
                        f"Sin PVP en tarifa {TARIFA_PERF}."))

        # ── ESTRUCTURA ─────────────────────────────────────────────────────────
        comps = comp_p_det.get(cod, {})

        # Mecanizado (col P Excel): debe estar en estructura del grupo 0
        mec_cp = str(cp.get("mecanizado", "") or "").strip()
        if mec_cp and grp_p == 0:
            if mec_cp not in comps:
                inc.append((cod, "ESTRUCTURA", "ERROR",
                            f"Mecanizado {mec_cp} debe estar en estructura del grupo 0."))

        # Estructura F2 grupo 0: debe llevar ENSAMBLADO si aplica
        if fam_p == 2 and grp_p == 0:
            tiene_ensamblado = any(k.upper().startswith(ENSAMBLADO_PREFIJO) for k in comps)
            if tiene_ensamblado:
                # Validar coste ENSAMBLADO
                for comp_cod, comp_row in comps.items():
                    if comp_cod.upper().startswith(ENSAMBLADO_PREFIJO):
                        coste_ens_art = num(comp_row.get("COSTEM"))
                        if coste_ens_art is not None and abs(coste_ens_art - ensamblado_coste_ref) > 0.01:
                            inc.append((cod, "ESTRUCTURA", "AVISO",
                                        f"Coste ENSAMBLADO {coste_ens_art} ≠ referencia {ensamblado_coste_ref}."))

        # Poliamidas: validar costes M/L
        for poli_key in ("poli1", "poli2"):
            poli_cod = str(cp.get(poli_key, "") or "").strip()
            cant_key = "cant_poli1" if poli_key == "poli1" else "cant_poli2"
            if poli_cod:
                coste_ml = num(coste_poli.get(poli_cod))
                if coste_ml is None:
                    inc.append((cod, "ESTRUCTURA", "AVISO",
                                f"Poliamida {poli_cod} no encontrada en tabla de costes."))

        # ── ALMACÉN (almacén 0) ────────────────────────────────────────────────
        sm_p = num(stockmin_p.get(cod))
        smax_p = num(stockmax_p.get(cod))

        def _jm_min_bool(val):
            if val is None: return False
            if isinstance(val, str): return val.strip().upper() == "SI"
            return num(val) and float(val) > 0

        if grp_p == 0:
            necesita_min = _jm_min_bool(_jm_min_bruto.get(base_p or cod))
        elif grp_p == 3:
            necesita_min = _jm_min_bool(_jm_min_blanco.get(base_p or cod))
        elif grp_p == 1:
            necesita_min = _jm_min_bool(_jm_min_plata.get(base_p or cod))
        else:
            necesita_min = False

        if necesita_min:
            if sm_p is None or sm_p <= 0:
                inc.append((cod, "STOCK_MINIMO", "ERROR",
                            "Excel JM indica stock mínimo requerido pero almacén tiene 0 o sin valor."))
            if smax_p is None or smax_p <= 0:
                inc.append((cod, "STOCK_MINIMO", "AVISO",
                            "Excel JM indica stock mínimo requerido pero stock máximo es 0."))
        else:
            if sm_p and sm_p > 0:
                inc.append((cod, "STOCK_MINIMO", "AVISO",
                            f"Stock mínimo {sm_p} definido pero Excel JM no indica mínimo para este grupo."))

        # Lote almacén (unidades internas = barras) convertido desde lote proveedor
        # F1 gr0 (proveedor en KG): lote_alm = lote_prov / (peso × ud_art)
        # Resto (proveedor en MT):  lote_alm = lote_prov / ud_art
        if grp_p == 0:
            lote_alm = num(lote_alm_p.get(cod))
            lote_prov2 = num(lote_pr_p.get(cod))
            if lote_prov2 is not None and lote_alm is not None and ud_art_p and ud_art_p > 0:
                if fam_p == 1 and peso_m and peso_m > 0:
                    lote_alm_esp = lote_prov2 / (peso_m * ud_art_p)
                else:
                    lote_alm_esp = lote_prov2 / ud_art_p
                if abs(lote_alm - lote_alm_esp) > 0.5:
                    inc.append((cod, "LOTE_ALMACEN", "ERROR",
                                f"Lote almacén {lote_alm} debe ser {lote_alm_esp:.1f} "
                                f"(lote prov {lote_prov2} / {'peso×ud' if fam_p==1 else 'ud_art'})."))

    # ── FIN PERFILES F1/F2 ─────────────────────────────────────────────────────

    # Orden de campos según pestañas de Geinfor (Maestro → Más → Mas2 → Caract. → Proveedor → Tarifa)
    _ORDEN_CAMPO = [
        # Maestro de artículos
        "FAMILIA", "CLASE", "DENOMINACION", "DENOMINACION_2", "ARTICULOBASE", "GRUPO",
        "TIPO", "TIPOARTMODULAR", "ESTADO", "MEDIDA", "COEFICIENTE", "UNIDADES_ARTICULO", "UNIDAD/TIPO",
        # Pestaña Más
        "PARTIDAARANCELARIA", "CONVERSION_M_KG",
        # Pestaña Mas2
        "GESTIONARVARIABLES", "ESSUBCONTRATACIO", "ESKIT_CHECK", "NOGESTIONASTOCK", "CALCULOCANTFORMULA",
        # Características
        "CARACTERISTICA_0",
        # Artículo proveedor
        "PROVEEDOR", "REF_PROVEEDOR", "COSTE_BRUTO", "COEF_PROVEEDOR", "PLAZO_ENTREGA",
        "LOTE_HABITUAL", "LOTE_ALMACEN", "UNIDAD_PROVEEDOR", "FORMULA_PROVEEDOR",
        # Tarifa / PVP
        "IVA_TARIFA", "UNIDAD_VENTA", "DIVISOR", "MARGEN", "PVP",
        # Proveedor - IVA
        "IVA_PROVEEDOR",
        # Artículo variable / asociado
        "ARTICULO_VARIABLE", "ARTICULO_ASOCIADO",
        # Excel JM
        "EXCEL_JM",
    ]
    _ord = {c: i for i, c in enumerate(_ORDEN_CAMPO)}

    out = pd.DataFrame(inc, columns=["CODIGO", "CAMPO", "NIVEL", "DETALLE"])
    out["_ord"] = out["CAMPO"].map(lambda c: _ord.get(c, 999))
    out = out.sort_values(["CODIGO", "_ord", "NIVEL"]).drop(columns="_ord").reset_index(drop=True)

    if articulo is not None:
        # Modo un artículo: SIEMPRE genera informe visual HTML (no toca el Excel).
        ruta = "INFORME_" + re.sub(r"[^A-Za-z0-9]", "_", articulo) + ".html"
        if len(sub) == 0 and len(sub_105) == 0 and len(sub_p) == 0:
            fila_m = m[m["CODIGO_ARTICULO"] == articulo]
            if len(fila_m) == 0:
                informe_mensaje(articulo, "Artículo no encontrado",
                                "Este código no existe en Geinfor. Revisa que esté bien escrito.", ruta)
                print(f"{articulo}: no encontrado. Informe -> {ruta}")
            else:
                famx = ni(fila_m.iloc[0].get("FAMILIAPRODUCTO"))
                informe_mensaje(articulo, "Familia no reconocida",
                                f"Familia de producto {famx}. El validador no cubre esta familia.", ruta, clase="bad")
                print(f"{articulo}: familia {famx} no reconocida. Informe -> {ruta}")
            try:
                import os
                os.startfile(os.path.abspath(ruta))
            except Exception:
                pass
            return
        inc_art = [t for t in inc if t[0] == articulo]
        if len(sub_p) > 0:
            fila_art = sub_p.iloc[0]
        elif len(sub) > 0:
            fila_art = sub.iloc[0]
        else:
            fila_art = sub_105.iloc[0]
        generar_informe_html(articulo, fila_art, inc_art, ruta)
        ne = sum(1 for t in inc_art if t[2] == "ERROR")
        na = sum(1 for t in inc_art if t[2] == "AVISO")
        print(f"{articulo}: {ne} error(es), {na} aviso(s). Informe -> {ruta}")
        try:
            import os
            os.startfile(os.path.abspath(ruta))  # abre el informe en el navegador (Windows)
        except Exception:
            pass
        return

    # Modo lote: informe visual HTML (principal) + Excel (para análisis en tabla).
    if desde is not None and hasta is not None:
        etiqueta = f"Altas creadas entre {desde.date()} y {hasta.date()}"
    elif desde is not None:
        etiqueta = f"Altas creadas desde {desde.date()}"
    else:
        etiqueta = "Todas las altas"
    ruta_html = "INFORME_ALTAS.html"
    sub_todo = pd.concat([sub, sub_105, sub_p], ignore_index=True)
    generar_informe_lote_html(sub_todo, out, etiqueta, ruta_html)

    destino = "INCIDENCIAS_FINAL.xlsx"
    try:
        out.to_excel(destino, index=False)
    except PermissionError:
        destino = "INCIDENCIAS_" + pd.Timestamp.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        out.to_excel(destino, index=False)
        print(f"(INCIDENCIAS_FINAL.xlsx estaba abierto; guardado en {destino})")

    print(f"Accesorios evaluados: {len(sub)} (+ {len(sub_105)} chapas/paneles familia 105, + {len(sub_p)} perfiles F1/F2)")
    print(f"Incidencias: {len(out)} (ERROR: {(out.NIVEL=='ERROR').sum()}, AVISO: {(out.NIVEL=='AVISO').sum()})")
    print(f"Informe visual -> {ruta_html}  |  Tabla -> {destino}")
    try:
        import os
        os.startfile(os.path.abspath(ruta_html))
    except Exception:
        pass


if __name__ == "__main__":
    main()
