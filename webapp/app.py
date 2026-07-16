#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys, os, json, subprocess, threading, io
from datetime import datetime, date
from flask import Flask, render_template, jsonify, request, Response, stream_with_context

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl

app = Flask(__name__)

EXCEL_GEN  = r'C:\Users\adrian\Desktop\Validador Articulos\GENERADOR ARTICULOS DIMAC.xlsx'
EXCEL_JM   = r'\\Dimac2023\datos\DATOS DIMAC\USERS\INTERCAMBIO\RAUL\xavi\PROCES DE COMPRES ADRIA\EXTRUSORES\MATRICES EXTRUSIÓ\20260628 Información y seguimiento Matrices nuevas.xlsx'

SHEETS_SEG = [
    ('SEGUIMIENTO MATRICES NUEVAS',      'matrices_nuevas'),
    ('SEGUIMIENTO MATRICES DUPLICADAS',  'matrices_dup'),
    ('SEGUIMIENTO ACCESORIOS NUEVOS',    'accesorios'),
    ('RELLENAR JOSE MARIA',              'jose_maria'),
]

SHEET_BY_KEY = {k: n for n, k in SHEETS_SEG}

ALERT_DAYS = 5

# ── helpers ──────────────────────────────────────────────────────────────────

def fmt_val(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y')
    return str(v)

def days_ago(v):
    if isinstance(v, datetime):
        return (datetime.now() - v).days
    if isinstance(v, date):
        return (date.today() - v).days
    return None

def load_sheet(path, sheet_name):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    headers = [fmt_val(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = [fmt_val(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any(v.strip() for v in row):
            rows.append(row)
    return headers, rows

def last_activity(row):
    """Return the most recent date found in a row."""
    best = None
    for v in row:
        if not v:
            continue
        for fmt in ('%d/%m/%Y',):
            try:
                d = datetime.strptime(v, fmt)
                if best is None or d > best:
                    best = d
            except ValueError:
                pass
    return best

def get_alerts():
    alerts = []
    for sheet_name, key in SHEETS_SEG:
        try:
            path = EXCEL_GEN if sheet_name in [n for n, _ in SHEETS_SEG[:3]] else EXCEL_JM
            # try gen excel first, fallback to JM
            try:
                wb = openpyxl.load_workbook(EXCEL_GEN, data_only=True)
                if sheet_name not in wb.sheetnames:
                    raise KeyError
                ws = wb[sheet_name]
            except Exception:
                wb = openpyxl.load_workbook(EXCEL_JM, data_only=True)
                ws = wb[sheet_name]

            headers = [fmt_val(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1):
                row = [fmt_val(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
                if not any(v.strip() for v in row):
                    continue
                ref = row[1] if len(row) > 1 else row[0]
                if not ref:
                    continue
                # check if finished (all steps OK/date, none NO)
                has_no = any(v.strip().upper() == 'NO' for v in row[2:])
                if not has_no:
                    continue
                last = last_activity(row)
                if last:
                    d = (datetime.now() - last).days
                    if d >= ALERT_DAYS:
                        # find first pending step
                        pending = next((headers[i] for i, v in enumerate(row) if v.strip().upper() == 'NO' and i < len(headers)), '—')
                        alerts.append({
                            'sheet': sheet_name,
                            'ref': ref,
                            'days': d,
                            'last': last.strftime('%d/%m/%Y'),
                            'pending': pending,
                        })
        except Exception as e:
            pass
    alerts.sort(key=lambda x: x['days'], reverse=True)
    return alerts

# ── routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def dashboard():
    alerts = get_alerts()
    # stats
    stats = {'alerts': len(alerts), 'sheets': len(SHEETS_SEG)}
    try:
        wb = openpyxl.load_workbook(EXCEL_GEN, data_only=True)
        ws = wb['PERFILES']
        listo = sum(1 for r in range(3, ws.max_row+1)
                    if str(ws.cell(r, 69).value or '').strip().upper() == 'LISTO')
        stats['perfiles_listo'] = listo
    except Exception:
        stats['perfiles_listo'] = '—'
    return render_template('dashboard.html', alerts=alerts, stats=stats)

@app.route('/seguimiento/<key>')
def seguimiento(key):
    sheet_name = SHEET_BY_KEY.get(key)
    if not sheet_name:
        return 'Not found', 404
    try:
        wb = openpyxl.load_workbook(EXCEL_GEN, data_only=True)
        if sheet_name in wb.sheetnames:
            headers, rows = load_sheet(EXCEL_GEN, sheet_name)
        else:
            headers, rows = load_sheet(EXCEL_JM, sheet_name)
    except Exception as e:
        headers, rows = [], []
    return render_template('seguimiento.html', sheet_name=sheet_name, key=key,
                           headers=headers, rows=rows, sheets=SHEETS_SEG, alert_days=ALERT_DAYS)

@app.route('/generador')
def generador():
    perfiles = []
    try:
        wb = openpyxl.load_workbook(EXCEL_GEN, data_only=True)
        ws = wb['PERFILES']
        for r in range(3, ws.max_row+1):
            ref = str(ws.cell(r, 1).value or '').strip()
            den = str(ws.cell(r, 3).value or '').strip()
            est = str(ws.cell(r, 69).value or '').strip().upper()
            if ref:
                perfiles.append({'ref': ref, 'den': den, 'estado': est})
    except Exception:
        pass
    return render_template('generador.html', perfiles=perfiles)

@app.route('/generador/simular', methods=['POST'])
def simular():
    ref = request.json.get('ref', '').strip().upper() or None
    def run():
        script = os.path.join(os.path.dirname(__file__), '..', 'generador_dimac.py')
        proc = subprocess.Popen(
            [sys.executable, '-c',
             f'''
import sys, io
sys.path.insert(0, r"C:\\Users\\adrian\\Desktop\\Validador Articulos")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import pyodbc, generador_dimac as g
rows = g.read_perfiles(filtro_ref={repr(ref)})
if not rows:
    print("Sin perfiles LISTO" + (" para " + {repr(ref)} if {repr(ref)} else ""))
else:
    cn = pyodbc.connect("DSN=GeinprodINGENRED", readonly=True)
    g.dry_run(rows, cn)
    cn.close()
'''],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            encoding='utf-8', errors='replace'
        )
        for line in proc.stdout:
            yield f"data: {json.dumps(line.rstrip())}\n\n"
        yield "data: __END__\n\n"
    return Response(stream_with_context(run()), mimetype='text/event-stream')

@app.route('/validador')
def validador():
    return render_template('validador.html')

@app.route('/validador/run', methods=['POST'])
def validador_run():
    modo = request.json.get('modo', 'todas')
    ref  = request.json.get('ref', '').strip().upper() or None
    def run():
        script = os.path.join(os.path.dirname(__file__), '..', 'validador_articulos_dimac.py')
        proc = subprocess.Popen(
            [sys.executable, script],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            encoding='utf-8', errors='replace',
            cwd=os.path.join(os.path.dirname(__file__), '..')
        )
        for line in proc.stdout:
            yield f"data: {json.dumps(line.rstrip())}\n\n"
        yield "data: __END__\n\n"
    return Response(stream_with_context(run()), mimetype='text/event-stream')

if __name__ == '__main__':
    app.run(debug=True, port=5000, use_reloader=False)
