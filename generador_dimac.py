#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GENERADOR ARTICULOS DIMAC - Phase 2
Lee GENERADOR ARTICULOS DIMAC.xlsx y crea articulos en Geinfor (DB2).
"""

import sys
import pyodbc
import openpyxl
from datetime import date

sys.stdout.reconfigure(encoding='utf-8')

# ============================================================
# CONFIG
# ============================================================

EXCEL_PATH   = r'C:\Users\adrian\Desktop\GENERADOR ARTICULOS DIMAC.xlsx'
DSN          = 'GeinprodINGENRED'
USUARIO      = 60    # CREADOPOR / MODIFICADOPOR en Geinfor
ESTADO_ART   = 70    # estado activo

CC_VEN_NAC   = '700000000'
CC_VEN_EXT   = '700021000'
CC_VEN_INTRA = '700031000'
CC_COMPRAS   = '600011001'

# ============================================================
# COLUMNAS Excel PERFILES (1-indexed)
# ============================================================

C_REF      = 1
C_FAMILIA  = 2
C_DEN      = 3
C_NORMA    = 5

C_G0 = 7;  C_G1 = 8;  C_G2 = 9;  C_G3 = 10
C_G4 = 11; C_G6 = 12; C_G7 = 13; C_G8 = 14; C_G9 = 15

C_PESO    = 16; C_LONG    = 17
C_PERIM_T = 18; C_PERIM_E = 19; C_CARAS = 20

C_POLI1 = 23; C_CANT_P1 = 24
C_POLI2 = 25; C_CANT_P2 = 26
C_COMP1 = 27; C_CANT_C1 = 28
C_COMP2 = 29; C_CANT_C2 = 30
C_MECAN = 31

C_PROVEEDOR  = 32; C_REF_PROV   = 33
C_LOTE_HAB   = 34; C_PLAZO_DIAS = 35; C_PARTIDA = 36

C_MIN_BT = 37; C_MIN_BL = 38; C_MIN_PL = 39

CST_BASE = 51   # cols 51-59: BT PL BR BL ASP STD ESP BI MAD
PVP_BASE = 60   # cols 60-68

C_ESTADO = 69

# ============================================================
# DEFINICION DE GRUPOS
# ============================================================
# (num, col_si, sufijo, denom2, art_acabado, idx_coste, write_cost)

# (grp_num, col_si, cod_sfx, sufijo_den, denom2, art_acabado, idx_coste, write_cost)
# cod_sfx: sufijo en el codigo de articulo (2 digitos).
# ATENCION: G4 ASP usa '05' (no '04') — patron confirmado en Geinfor con A000472805.
GRUPOS = [
    (0, C_G0, None,  'BT',  'BRUTO',              None,                0, True),
    (1, C_G1, '01',  'PL',  'PLATA MATE',         'PLATA MATE',        1, True),
    (2, C_G2, '02',  'BR',  'BRONCE',             'ANODIZADOSBRONCE',  2, False),
    (3, C_G3, '03',  'BL',  'BLANCO',             'BLANCO',            3, True),
    (4, C_G4, '05',  'ASP', 'ANODIZADO ESPECIAL', 'ANODIZADOESPECIAL', 4, False),
    (6, C_G6, '06',  'STD', 'RAL ESTANDAR',       'RALESTANDAR',       5, False),
    (7, C_G7, '07',  'ESP', 'RAL ESPECIAL',       'RALESPECIAL',       6, False),
    (8, C_G8, '08',  'BI',  'BICOLOR',            'BICOLOR',           7, False),
    (9, C_G9, '09',  'MAD', 'MADERA',             'MADERA',            8, False),
]

GRUPOS_FORMULA_40 = {1, 3}   # acabado con cant=PERIM×LONG, recalcular=1

# ============================================================
# TARIFAS  (cod, base, mult, ud_venta, iva)
# ============================================================

TARIFAS = [
    (40,    'T40',   1.00, 4, 21),
    (4010,  'T40',   0.90, 4, 21),
    (4015,  'T40',   0.85, 4, 21),
    (4020,  'T40',   0.80, 4, 21),
    (40000, 'COSTE', 1.15, 9, 21),
    (50000, 'COSTE', 1.19, 9, 21),
    (60000, 'COSTE', 1.25, 9, 21),
    (90000, 'COSTE', 1.00, 9, 21),
    (40100, 'T40',   1.00, 4, 0),
    (40103, 'T40',   1.03, 4, 0),
    (40106, 'T40',   1.06, 4, 0),
    (40109, 'T40',   1.09, 4, 0),
    (40112, 'T40',   1.12, 4, 0),
    (40115, 'T40',   1.15, 4, 0),
    (40118, 'T40',   1.18, 4, 0),
    (40121, 'T40',   1.21, 4, 0),
    (40133, 'T40',   1.33, 4, 0),
    (40136, 'T40',   1.36, 4, 0),
]

# ============================================================
# HELPERS
# ============================================================

def cel(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return v

def s(v): return str(v).strip() if v is not None else ''
def n(v): return v if v is not None else 0
def f(v): return float(n(v))
def i(v):
    try: return int(n(v))
    except (ValueError, TypeError): return 0

def art_code(ref, cod_sfx):
    return ref.strip() if cod_sfx is None else f"{ref.strip()}{cod_sfx}"

def get_tipo(familia, grp):
    if grp == 0: return 1 if familia == 1 else 16
    if grp in (1, 3): return 60
    return 92

def get_tipomod(tipo, grp):
    if tipo in (1, 16, 60): return 0
    return 0 if grp == 8 else 1

def get_coef(familia, grp, peso, long_):
    if familia == 1 and grp == 0:
        return round(1.0 / (peso * long_), 5) if peso * long_ > 0 else 0
    return round(1.0 / long_, 5) if long_ > 0 else 0

def get_ud_ext(familia, grp):
    return 2 if (familia == 1 and grp == 0) else 4

def perim_str(val):
    return str(round(val, 3)).replace('.', ',')

def base_den(den_raw):
    """Quita sufijo de grupo si el usuario lo puso en la denominacion."""
    den = den_raw.strip()
    for _, _, _, sf, *_ in GRUPOS:
        if sf and den.upper().endswith(' ' + sf):
            den = den[:-(len(sf)+1)].strip()
            break
    return den

# ============================================================
# LEER EXCEL
# ============================================================

def lookup_proveedor(val_raw):
    """Devuelve (codigo_int, ref_texto) dado un valor de la celda PROVEEDOR.
    Acepta entero directo o texto (alias/nombre) que busca en MAESTRO_PROVEEDORES."""
    val = s(val_raw)
    if not val:
        return 0, ''
    try:
        return int(val), val
    except ValueError:
        pass
    # Buscar por alias o razon_social
    try:
        cn = pyodbc.connect(f'DSN={DSN}', readonly=True)
        cur = cn.cursor()
        cur.execute("""SELECT CODIGO_PROVEEDOR FROM DB2ADMIN.MAESTRO_PROVEEDORES
                       WHERE UPPER(ALIAS)=? OR UPPER(RAZON_SOCIAL) LIKE ?""",
                    val.upper(), f'%{val.upper()}%')
        row = cur.fetchone()
        cn.close()
        if row:
            return int(row[0]), val
    except Exception:
        pass
    print(f"  [AVISO] Proveedor '{val}' no encontrado en Geinfor. Asignado como 0.")
    return 0, val


def read_perfiles(filtro_ref=None):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['PERFILES']
    rows = []

    for rn in range(3, ws.max_row + 1):
        ref = s(cel(ws, rn, C_REF))
        if not ref:
            continue
        estado = s(cel(ws, rn, C_ESTADO))

        if filtro_ref:
            if ref.upper() != filtro_ref.upper():
                continue
        else:
            if estado.upper() != 'LISTO':
                continue

        grupos_si = set()
        for grp_num, col_si, cod_sfx, *_ in GRUPOS:
            if s(cel(ws, rn, col_si)).upper() == 'SI':
                grupos_si.add(grp_num)

        if not grupos_si:
            continue

        rd = {
            'ref':      ref,
            'familia':  i(cel(ws, rn, C_FAMILIA)) or 1,
            'den':      s(cel(ws, rn, C_DEN)),
            'norma':    s(cel(ws, rn, C_NORMA)) or 'EU',
            'peso':     f(cel(ws, rn, C_PESO)),
            'long':     f(cel(ws, rn, C_LONG)),
            'perim_t':  f(cel(ws, rn, C_PERIM_T)),
            'perim_e':  f(cel(ws, rn, C_PERIM_E)),
            'caras':    i(cel(ws, rn, C_CARAS)),
            'poli1':    s(cel(ws, rn, C_POLI1)),  'cant_p1': f(cel(ws, rn, C_CANT_P1)),
            'poli2':    s(cel(ws, rn, C_POLI2)),  'cant_p2': f(cel(ws, rn, C_CANT_P2)),
            'comp1':    s(cel(ws, rn, C_COMP1)),  'cant_c1': f(cel(ws, rn, C_CANT_C1)),
            'comp2':    s(cel(ws, rn, C_COMP2)),  'cant_c2': f(cel(ws, rn, C_CANT_C2)),
            'mecan':    s(cel(ws, rn, C_MECAN)),
            'prov_int': lookup_proveedor(cel(ws, rn, C_PROVEEDOR))[0],
            'ref_prov': s(cel(ws, rn, C_REF_PROV)) or lookup_proveedor(cel(ws, rn, C_PROVEEDOR))[1],
            'lote':     f(cel(ws, rn, C_LOTE_HAB)) or 1.0,
            'plazo':    i(cel(ws, rn, C_PLAZO_DIAS)),
            'partida':  i(cel(ws, rn, C_PARTIDA)) or 1,
            'min_bt':   f(cel(ws, rn, C_MIN_BT)),
            'min_bl':   f(cel(ws, rn, C_MIN_BL)),
            'min_pl':   f(cel(ws, rn, C_MIN_PL)),
            'costes':   [f(cel(ws, rn, CST_BASE + k)) for k in range(9)],
            'pvps':     [f(cel(ws, rn, PVP_BASE + k)) for k in range(9)],
            'grupos_si': grupos_si,
        }
        rows.append(rd)

    return rows

# ============================================================
# CONSTRUIR REGISTRO DE ARTICULO
# ============================================================

def build_art(rd, grp_num, cod_sfx, sufijo, denom2, acabado, idx_coste, write_cost):
    long_  = rd['long']
    peso   = rd['peso']
    fam    = rd['familia']
    cod    = art_code(rd['ref'], cod_sfx)
    tipo   = get_tipo(fam, grp_num)
    coef   = get_coef(fam, grp_num, peso, long_)
    ud_ext = get_ud_ext(fam, grp_num)
    coste_m   = rd['costes'][idx_coste]
    coste_bar = round(coste_m * long_, 6) if write_cost else 0.0
    pvp_t40_m = rd['pvps'][idx_coste]

    den_base = base_den(rd['den'])
    denominacion = f"{den_base} {sufijo}"[:100]

    return {
        'codigo':          cod,
        'denominacion':    denominacion,
        'denominacion_2':  denom2,
        'clase':           51,
        'tipo':            tipo,
        'tipomod':         get_tipomod(tipo, grp_num),
        'norma':           rd['norma'],
        'coste':           coste_bar,
        'grupo':           grp_num,
        'familia':         fam,
        'ud_int':          9,
        'ud_ext':          ud_ext,
        'unidades_art':    long_,
        'coef':            coef,
        'conv_m_kg':       peso,
        'peso_art':        peso if grp_num != 0 else 0.0,
        'peso_bulto':      round(peso * long_, 4) if grp_num != 0 else 0.0,
        'partida':         rd['partida'],
        'prov_int':        rd['prov_int'],
        'ref_prov':        rd['ref_prov'],
        # privados para calculos
        '_rd':             rd,
        '_grp':            grp_num,
        '_acabado':        acabado,
        '_pvp_t40_m':      pvp_t40_m,
        '_coste_m':        coste_m,
    }

def arts_for_row(rd):
    arts = []
    for grp_num, col_si, cod_sfx, sufijo, denom2, acabado, idx_coste, write_cost in GRUPOS:
        if grp_num in rd['grupos_si']:
            arts.append(build_art(rd, grp_num, cod_sfx, sufijo, denom2, acabado, idx_coste, write_cost))
    return arts

# ============================================================
# CALCULO PVP TARIFA
# ============================================================

def calc_pvp(art, tar):
    cod_tar, base, mult, ud_venta, iva = tar
    long_     = art['unidades_art']
    pvp_t40_m = art['_pvp_t40_m']
    coste_m   = art['_coste_m']

    pvp  = round(pvp_t40_m * mult, 4) if base == 'T40' else round(coste_m * long_ * mult, 4)
    coef = round(1.0 / long_, 5) if ud_venta == 4 and long_ > 0 else 1.0
    return pvp, float(iva), coef, ud_venta

# ============================================================
# DRY-RUN
# ============================================================

def dry_run(rows, cn):
    import io, os
    buf = io.StringIO()

    def out(msg=''):
        print(msg)
        buf.write(msg + '\n')

    cur = cn.cursor()
    total_new = 0; total_skip = 0

    for rd in rows:
        out(f"\n{'='*65}")
        out(f"  REF={rd['ref']}  F{rd['familia']}  grupos={sorted(rd['grupos_si'])}")
        out(f"  '{rd['den']}'  peso={rd['peso']}  long={rd['long']}  perim={rd['perim_t']}")

        for art in arts_for_row(rd):
            cod = art['codigo']
            cur.execute("SELECT DENOMINACION FROM DB2ADMIN.MAESTRO_DE_ARTICULOS WHERE CODIGO_ARTICULO=?", cod)
            existe = cur.fetchone()
            if existe:
                out(f"  [SKIP] {cod} — ya existe ({existe[0]})")
                total_skip += 1
                continue

            grp = art['_grp']
            rd2 = art['_rd']
            long_ = art['unidades_art']

            out(f"\n  [CREAR] {cod}  '{art['denominacion']}'")
            out(f"    tipo={art['tipo']} grupo={grp} fam={art['familia']} tipomod={art['tipomod']}")
            out(f"    coste={art['coste']:.4f} €/barra  coef={art['coef']}  ud_ext={art['ud_ext']}")

            if grp == 0 and art['familia'] == 1:
                out(f"    estructura: (ninguna — F1 G0)")
            elif grp == 0 and art['familia'] == 2:
                out(f"    estructura: F2 G0 — poli/comp desde Excel + ENSAMBLADO")
            else:
                if grp in GRUPOS_FORMULA_40:
                    cant_acab = round(rd2['perim_t'] * long_, 4)
                    out(f"    estructura: {rd2['ref']}×1  +  {art['_acabado']}×{cant_acab} (formula=40)")
                else:
                    out(f"    estructura: {rd2['ref']}×1  +  {art['_acabado']}×1")

            out(f"    carac: PLANO={rd2['ref']}.JPG  CARAS={rd2['caras']}  PERIM_T={rd2['perim_t']}  PERIM_E={rd2['perim_e']}")

            pvp40, _, _, _ = calc_pvp(art, TARIFAS[0])
            pvp40k, _, _, _ = calc_pvp(art, next(t for t in TARIFAS if t[0] == 40000))
            out(f"    tarifas: T40={pvp40} €/m  |  40000={pvp40k} €/barra  ({len(TARIFAS)} total)")

            if grp == 0:
                piva = 0.0 if art['norma'] == 'TURQUIA' else 21.0
                out(f"    proveedor: {rd2['prov_int']} ({rd2['ref_prov']}) IVA={piva}  lote={rd2['lote']}  plazo={rd2['plazo']}")

            min_val = {0: rd2['min_bt'], 1: rd2['min_pl'], 3: rd2['min_bl']}.get(grp, 0)
            out(f"    almacen: ORIGEN=0 ALMACEN=0 STOCKMIN={min_val}")

            total_new += 1

    out(f"\n{'='*65}")
    out(f"RESUMEN SIMULACION: {total_new} a crear  |  {total_skip} ya existen")

    log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'simulacion_log.txt')
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write(buf.getvalue())
    print(f"\nLog guardado en: {log_path}")

    cur.close()

# ============================================================
# INSERTS
# ============================================================

def ins_maestro(cur, art):
    cur.execute("""
        INSERT INTO DB2ADMIN.MAESTRO_DE_ARTICULOS (
            CODIGO_ARTICULO, DENOMINACION, DENOMINACION_2,
            CLASE, TIPO, TIPOARTMODULAR, NORMA, COSTE, GRUPO, FAMILIAPRODUCTO,
            MEDIDA, UNIDAD_INTERNA, UNIDAD_EXTERNA, UNIDADES_ARTICULO, COEFICIENTE,
            CONVERSION_M_KG, PESO_ARTICULO, PESO_BULTO, UNIDADBULTO,
            PROVEEDOR, CODIGO_PROVEEDOR, ESTADO,
            CODIGO_CONTABLE, CC_EXTRANJERO, CC_INTRACOMUNITARIA,
            CC_COMPRAS_NAC, CC_COMPRAS_INT, CC_COMPRAS_EXP,
            PARTIDAARANCELARIA, CALCULOCANTFORMULA, GESTIONARVARIABLES,
            COEFPARTARANCELARIA, CATEGORIAGLOBAL, CATEGORIACALCULO,
            PUBLICARWEB, SECUENCIAS, MOSTRAR_INF_MRP,
            CREADOPOR, MODIFICADOPOR, FECHAALTA, FECHACAMBIOESTADO
        ) VALUES (
            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,
            CURRENT DATE, CURRENT DATE
        )
    """,
    art['codigo'], art['denominacion'], art['denominacion_2'],
    art['clase'], art['tipo'], art['tipomod'], art['norma'], art['coste'],
    art['grupo'], art['familia'],
    1, art['ud_int'], art['ud_ext'], art['unidades_art'], art['coef'],
    art['conv_m_kg'], art['peso_art'], art['peso_bulto'], 1,
    art['prov_int'], art['ref_prov'], ESTADO_ART,
    CC_VEN_NAC, CC_VEN_EXT, CC_VEN_INTRA,
    CC_COMPRAS, CC_COMPRAS, CC_COMPRAS,
    art['partida'], 1, 1,
    1.0, 5, 5,
    1, 1, 1,
    USUARIO, USUARIO
    )


def ins_estructura(cur, art):
    grp  = art['_grp']
    rd   = art['_rd']
    cod  = art['codigo']
    long_ = art['unidades_art']
    fam  = art['familia']

    if grp == 0 and fam == 1:
        return  # F1 G0: sin estructura

    def ins(comp, cant, formula=0, recalc=0):
        cur.execute("""
            INSERT INTO DB2ADMIN.MAESTRO_ESTRUCTURAS
            (ARTSUPERIOR, OPCIONESTRUCTURA, ARTCOMPONENTE, CANTIDAD, FORMULA, RECALCULARFORMULA, ESBAJA)
            VALUES (?, 0, ?, ?, ?, ?, 0)
        """, cod, comp, cant, formula, recalc)

    if grp == 0 and fam == 2:
        # F2 G0: poliamidas + componentes (ENSAMBLADO pendiente de verificar)
        print(f"    [WARNING] F2 G0 estructura incompleta — falta codigo ENSAMBLADO")
        for comp, cant in [(rd['poli1'], rd['cant_p1']), (rd['poli2'], rd['cant_p2']),
                           (rd['comp1'], rd['cant_c1']), (rd['comp2'], rd['cant_c2'])]:
            if comp:
                ins(comp, cant)
        if rd['mecan']:
            ins(rd['mecan'], 1.0)
        return

    # G1-G9: base + acabado
    ins(rd['ref'], 1.0)

    if grp in GRUPOS_FORMULA_40:
        cant_acab = round(rd['perim_t'] * long_, 4)
        ins(art['_acabado'], cant_acab, formula=40, recalc=1)
    else:
        ins(art['_acabado'], 1.0)

    # F2 G1-G9: también polis y comps
    if fam == 2:
        for comp, cant in [(rd['poli1'], rd['cant_p1']), (rd['poli2'], rd['cant_p2']),
                           (rd['comp1'], rd['cant_c1']), (rd['comp2'], rd['cant_c2'])]:
            if comp:
                ins(comp, cant)


def ins_tarifas(cur, art):
    for tar in TARIFAS:
        pvp, iva, coef, ud = calc_pvp(art, tar)
        cur.execute("""
            INSERT INTO DB2ADMIN.ARTICULOS_TARIFABLES
            (ARTICULO, TARIFA, PRECIO_VENTA_PUBLICO, IVA, COEFICIENTE, UNIDADVENTA)
            VALUES (?, ?, ?, ?, ?, ?)
        """, art['codigo'], tar[0], pvp, iva, coef, ud)


def ins_caract(cur, art):
    rd  = art['_rd']
    cod = art['codigo']
    ref = rd['ref']
    for carac_id, valor in [
        (0,   f"X:\\IMAGENES GEINPROD\\{ref}.JPG"),
        (2,   f"X:\\IMAGENES GEINPROD\\PLANO ORIGINAL THERMIA\\{ref}.pdf"),
        (100, str(rd['caras'])),
        (200, perim_str(rd['perim_t'])),
        (201, perim_str(rd['perim_e'])),
    ]:
        cur.execute(
            "INSERT INTO DB2ADMIN.TCARAC_ART (ARTICULO, CARACTERISTICA, VALOR) VALUES (?,?,?)",
            cod, carac_id, valor
        )


def ins_variable(cur, art):
    cur.execute(
        "INSERT INTO DB2ADMIN.ARTICULOVARIABLE (ARTICULO, VARIABLE, POSICION) VALUES (?, 10, 0)",
        art['codigo']
    )


def ins_proveedor(cur, art):
    if art['_grp'] != 0:
        return
    rd   = art['_rd']
    piva = 0.0 if art['norma'] == 'TURQUIA' else 21.0
    cur.execute("""
        INSERT INTO DB2ADMIN.ARTICULOS_PROVEEDOR
        (ARTICULO, PROVEEDOR, CODIGO_PROVEEDOR, PRINCIPAL, PIVA,
         LOTEHABITUAL, PLAZOENTREGA, UNIDADEXTERNA, COEFICIENTE,
         NUMERO_DECIMALES, CANTMULTIPLO, APROXIMARMULTIPLO)
        VALUES (?, ?, ?, 1, ?, ?, ?, ?, ?, 2, 1.0, 1)
    """,
    art['codigo'], rd['prov_int'], rd['ref_prov'], piva,
    rd['lote'], rd['plazo'], art['ud_ext'], art['coef']
    )


def ins_almacen(cur, art):
    grp = art['_grp']
    rd  = art['_rd']
    min_val = {0: rd['min_bt'], 1: rd['min_pl'], 3: rd['min_bl']}.get(grp, 0.0)
    cur.execute("""
        INSERT INTO DB2ADMIN.ALMACEN_ARTICULOS
        (ARTICULO, ORIGEN, ALMACEN, STOCKMINIMO, STOCKMAXIMO)
        VALUES (?, 0, 0, ?, 0)
    """, art['codigo'], min_val)

# ============================================================
# GENERACION
# ============================================================

def generate(rows, cn):
    cur = cn.cursor()
    log = []
    ok = skip = err = 0

    for rd in rows:
        for art in arts_for_row(rd):
            cod = art['codigo']
            grp = art['_grp']

            cur.execute("SELECT 1 FROM DB2ADMIN.MAESTRO_DE_ARTICULOS WHERE CODIGO_ARTICULO=?", cod)
            if cur.fetchone():
                print(f"  [SKIP] {cod}")
                log.append((cod, grp, 'SKIP', 'ya existe'))
                skip += 1
                continue

            print(f"  [GEN]  {cod}  {art['denominacion']}", end='  ')
            try:
                ins_maestro(cur, art)
                ins_estructura(cur, art)
                ins_tarifas(cur, art)
                ins_caract(cur, art)
                ins_variable(cur, art)
                ins_proveedor(cur, art)
                ins_almacen(cur, art)
                cn.commit()
                print("OK")
                log.append((cod, grp, 'OK', ''))
                ok += 1
            except Exception as e:
                cn.rollback()
                msg = str(e)[:250]
                print(f"ERROR: {msg}")
                log.append((cod, grp, 'ERROR', msg))
                err += 1

    cur.close()
    print(f"\n{'='*65}")
    print(f"RESULTADO: {ok} creados | {skip} skip | {err} errores")
    return log


def write_seguimiento(log):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['SEGUIMIENTO']
    rn = 2
    while ws.cell(row=rn, column=1).value:
        rn += 1
    hoy = date.today().isoformat()
    for cod, grp, resultado, msg in log:
        ws.cell(row=rn, column=1).value = hoy
        ws.cell(row=rn, column=2).value = cod
        ws.cell(row=rn, column=3).value = grp
        ws.cell(row=rn, column=4).value = resultado
        ws.cell(row=rn, column=5).value = msg
        rn += 1
    wb.save(EXCEL_PATH)
    print(f"Log guardado en SEGUIMIENTO ({len(log)} entradas)")

# ============================================================
# MENU
# ============================================================

def menu():
    print("\n" + "="*65)
    print("  GENERADOR ARTICULOS DIMAC  —  Phase 2")
    print("="*65)
    print("  1. Simular todos los articulos con ESTADO=LISTO")
    print("  2. Simular una referencia concreta")
    print("  3. GENERAR todos los articulos con ESTADO=LISTO")
    print("  4. GENERAR una referencia concreta")
    print("  5. Salir")
    print("="*65)
    return input("  Opcion: ").strip()


def main():
    cn_ro = pyodbc.connect(f'DSN={DSN}', readonly=True)

    while True:
        op = menu()
        if op == '5':
            break

        if op not in ('1','2','3','4'):
            print("  Opcion no valida.")
            continue

        dry   = op in ('1','2')
        filtro = None
        if op in ('2','4'):
            filtro = input("  Referencia (ej. A0004728): ").strip().upper()

        print(f"\nLeyendo Excel...")
        rows = read_perfiles(filtro)

        if not rows:
            print("  Sin resultados. Verifica ESTADO=LISTO o la referencia.")
            continue

        print(f"  {len(rows)} perfil(es) encontrado(s).")

        if dry:
            print("\n--- MODO SIMULACION (no se escribe nada) ---")
            dry_run(rows, cn_ro)
        else:
            confirm = input(f"\n  Confirmar GENERACION? Escribe 'si' para continuar: ").strip().lower()
            if confirm != 'si':
                print("  Cancelado.")
                continue
            cn_rw = pyodbc.connect(f'DSN={DSN}')
            try:
                log = generate(rows, cn_rw)
                write_seguimiento(log)
            finally:
                cn_rw.close()

    cn_ro.close()
    print("Saliendo.")


if __name__ == '__main__':
    main()
